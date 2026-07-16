import hashlib
import html
import json
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from pandas.tseries.holiday import USFederalHolidayCalendar
from pandas.tseries.offsets import CustomBusinessDay

CUSTOMER_EXPORT_VERSION = "Customer export v12"
APP_CACHE_VERSION = "inventory-logic-v29-all-report-skus"
WAREHOUSE_BUSINESS_DAY = CustomBusinessDay(calendar=USFederalHolidayCalendar())


st.set_page_config(
    page_title="Inventory Shortage Dashboard",
    page_icon="▦",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    r"""
    <style>
        :root {
            --win-accent: #0067c0;
            --win-accent-hover: #005a9e;
            --win-accent-pressed: #004578;
            --win-accent-soft: rgba(0, 103, 192, .10);
            --win-text: #1f1f1f;
            --win-text-secondary: #5d5d5d;
            --win-text-tertiary: #737373;
            --win-bg: #f3f3f3;
            --win-surface: rgba(255, 255, 255, .88);
            --win-surface-solid: #ffffff;
            --win-surface-subtle: rgba(249, 249, 249, .82);
            --win-border: rgba(0, 0, 0, .0578);
            --win-border-strong: rgba(0, 0, 0, .12);
            --win-shadow-card: 0 1px 2px rgba(0,0,0,.04), 0 4px 14px rgba(0,0,0,.055);
            --win-shadow-flyout: 0 8px 28px rgba(0,0,0,.12), 0 2px 8px rgba(0,0,0,.06);
            --win-radius-sm: 6px;
            --win-radius-md: 8px;
            --win-radius-lg: 12px;
            --win-radius-xl: 16px;
            --win-ease: cubic-bezier(.1,.9,.2,1);
            --win-fast: 120ms;
            --win-normal: 180ms;
            --layout-top: 6px;
            --layout-x: 24px;
            --sidebar-width: 292px;
            --sidebar-control-top: 14px;
            --sidebar-control-size: 40px;
        }

        html { scroll-behavior: smooth; }
        body, .stApp {
            font-family: "Segoe UI Variable Text", "Segoe UI", system-ui, -apple-system, BlinkMacSystemFont, sans-serif;
            color: var(--win-text);
            background:
                radial-gradient(900px 440px at 82% -10%, rgba(105, 180, 255, .18), transparent 64%),
                radial-gradient(700px 360px at 10% 0%, rgba(188, 222, 255, .13), transparent 62%),
                var(--win-bg);
        }
        .stApp { min-height: 100vh; }
        [data-testid="stHeader"] {
            display: block !important;
            visibility: visible !important;
            position: fixed !important;
            top: 0 !important;
            left: 0 !important;
            right: 0 !important;
            height: 54px !important;
            min-height: 54px !important;
            overflow: visible !important;
            background: transparent !important;
            border: 0 !important;
            box-shadow: none !important;
            pointer-events: auto !important;
            z-index: 999990 !important;
        }
        [data-testid="stHeader"] > div {
            overflow: visible !important;
            background: transparent !important;
            pointer-events: auto !important;
        }
        footer,
        #MainMenu,
        [data-testid="stAppDeployButton"],
        [data-testid="stMainMenu"],
        [data-testid="stStatusWidget"],
        [data-testid="stDecoration"],
        [data-testid="stViewerBadge"],
        .stDeployButton,
        .viewerBadge_container__1QSob,
        [class*="viewerBadge"],
        [class*="ViewerBadge"],
        [data-testid="stHeader"] a[href*="github.com"],
        [data-testid="stHeader"] button[aria-label*="share" i],
        [data-testid="stHeader"] button[title*="share" i],
        [data-testid="stHeader"] button[aria-label*="edit" i],
        [data-testid="stHeader"] button[title*="edit" i],
        [data-testid="stHeader"] button[aria-label*="favorite" i],
        [data-testid="stHeader"] button[title*="favorite" i] {
            display: none !important;
        }
        [data-testid="stToolbar"],
        [data-testid="stHeaderActionElements"] {
            display: flex !important;
            visibility: visible !important;
            overflow: visible !important;
            background: transparent !important;
            pointer-events: auto !important;
        }
        [data-testid="stSidebarCollapsedControl"],
        [data-testid="collapsedControl"] {
            display: flex !important;
            visibility: visible !important;
            position: fixed !important;
            top: var(--sidebar-control-top) !important;
            left: 14px !important;
            z-index: 2147483646 !important;
            width: 108px !important;
            height: var(--sidebar-control-size) !important;
            min-width: 108px !important;
            min-height: var(--sidebar-control-size) !important;
            margin: 0 !important;
            padding: 0 !important;
            overflow: visible !important;
            border: 0 !important;
            background: transparent !important;
            pointer-events: auto !important;
            opacity: 1 !important;
        }
        [data-testid="stSidebarCollapsedControl"] button,
        [data-testid="collapsedControl"] button,
        button[aria-label*="open sidebar" i],
        button[title*="open sidebar" i] {
            display: inline-flex !important;
            visibility: visible !important;
            position: fixed !important;
            top: var(--sidebar-control-top) !important;
            left: 14px !important;
            z-index: 2147483647 !important;
            align-items: center !important;
            justify-content: flex-start !important;
            gap: 9px !important;
            width: 108px !important;
            height: var(--sidebar-control-size) !important;
            min-width: 108px !important;
            min-height: var(--sidebar-control-size) !important;
            margin: 0 !important;
            padding: 0 14px !important;
            color: #ffffff !important;
            font-family: "Segoe UI Variable Text", "Segoe UI", sans-serif !important;
            font-size: 12px !important;
            font-weight: 650 !important;
            line-height: 1 !important;
            white-space: nowrap !important;
            border: 1px solid rgba(0, 70, 130, .36) !important;
            border-radius: 10px !important;
            background: linear-gradient(180deg, #0a73c9 0%, #0067c0 100%) !important;
            box-shadow: 0 2px 5px rgba(0, 77, 138, .18), 0 8px 22px rgba(0, 77, 138, .22) !important;
            pointer-events: auto !important;
            touch-action: manipulation !important;
            cursor: pointer !important;
            opacity: 1 !important;
            transform: translateZ(0) !important;
            transition: background var(--win-fast) ease, box-shadow var(--win-normal) var(--win-ease), transform var(--win-fast) ease !important;
        }
        [data-testid="stSidebarCollapsedControl"] button svg,
        [data-testid="collapsedControl"] button svg,
        button[aria-label*="open sidebar" i] svg,
        button[title*="open sidebar" i] svg {
            display: none !important;
        }
        [data-testid="stSidebarCollapsedControl"] button::before,
        [data-testid="collapsedControl"] button::before,
        button[aria-label*="open sidebar" i]::before,
        button[title*="open sidebar" i]::before {
            content: "\E700";
            display: inline-flex;
            align-items: center;
            justify-content: center;
            width: 16px;
            height: 16px;
            color: #ffffff;
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            font-size: 15px;
            font-weight: 400;
            flex: 0 0 auto;
            pointer-events: none;
        }
        [data-testid="stSidebarCollapsedControl"] button::after,
        [data-testid="collapsedControl"] button::after,
        button[aria-label*="open sidebar" i]::after,
        button[title*="open sidebar" i]::after {
            content: "Filters";
            color: #ffffff;
            font-family: "Segoe UI Variable Text", "Segoe UI", sans-serif;
            font-size: 12px;
            font-weight: 650;
            letter-spacing: .01em;
            pointer-events: none;
        }
        [data-testid="stSidebarCollapsedControl"] button:hover,
        [data-testid="collapsedControl"] button:hover,
        button[aria-label*="open sidebar" i]:hover,
        button[title*="open sidebar" i]:hover {
            background: linear-gradient(180deg, #087bd9 0%, #005fae 100%) !important;
            box-shadow: 0 3px 7px rgba(0, 77, 138, .22), 0 11px 28px rgba(0, 77, 138, .27) !important;
            transform: translateY(-1px) !important;
        }
        [data-testid="stSidebarCollapsedControl"] button:active,
        [data-testid="collapsedControl"] button:active,
        button[aria-label*="open sidebar" i]:active,
        button[title*="open sidebar" i]:active {
            background: var(--win-accent-pressed) !important;
            box-shadow: 0 1px 4px rgba(0, 77, 138, .22) !important;
            transform: translateY(0) scale(.985) !important;
        }
        [data-testid="stSidebarCollapsedControl"] button:focus-visible,
        [data-testid="collapsedControl"] button:focus-visible,
        button[aria-label*="open sidebar" i]:focus-visible,
        button[title*="open sidebar" i]:focus-visible {
            outline: 2px solid rgba(0, 103, 192, .72) !important;
            outline-offset: 3px !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] {
            display: flex !important;
            visibility: visible !important;
            position: absolute !important;
            top: var(--sidebar-control-top) !important;
            right: 12px !important;
            z-index: 30 !important;
            width: var(--sidebar-control-size) !important;
            height: var(--sidebar-control-size) !important;
            pointer-events: auto !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button {
            display: inline-flex !important;
            align-items: center !important;
            justify-content: center !important;
            width: 100% !important;
            height: 100% !important;
            min-width: 0 !important;
            min-height: 0 !important;
            padding: 0 !important;
            color: var(--win-text-secondary) !important;
            border: 1px solid var(--win-border-strong) !important;
            border-radius: 10px !important;
            background: rgba(255,255,255,.82) !important;
            box-shadow: 0 1px 2px rgba(0,0,0,.04), 0 4px 12px rgba(0,0,0,.07) !important;
            pointer-events: auto !important;
            transition: color var(--win-fast) ease, background var(--win-fast) ease, border-color var(--win-fast) ease, box-shadow var(--win-fast) ease, transform var(--win-fast) ease !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button:hover {
            color: var(--win-text) !important;
            background: #ffffff !important;
            border-color: rgba(0,103,192,.28) !important;
            box-shadow: 0 2px 7px rgba(0,0,0,.10) !important;
            transform: translateY(-1px);
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button:active {
            background: #f4f4f4 !important;
            transform: translateY(0) scale(.97);
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"] button:focus-visible {
            outline: 2px solid rgba(0,103,192,.62) !important;
            outline-offset: 2px !important;
        }
        [data-testid="stAppViewContainer"] > .main,
        [data-testid="stMain"] {
            padding-top: 54px !important;
            margin-top: 0 !important;
        }
        .stApp:has([data-testid="stSidebarCollapsedControl"]) .main .block-container,
        .stApp:has([data-testid="collapsedControl"]) .main .block-container {
            padding-left: 138px !important;
        }

        .main .block-container {
            max-width: 1660px;
            padding: var(--layout-top) var(--layout-x) 36px;
            animation: pageReveal .36s var(--win-ease) both;
        }
        @keyframes pageReveal {
            from { opacity: 0; transform: translateY(5px); }
            to { opacity: 1; transform: translateY(0); }
        }

        /* Sidebar: same top rhythm and surface language as the workspace */
        section[data-testid="stSidebar"] {
            width: var(--sidebar-width) !important;
            min-width: var(--sidebar-width) !important;
            background: rgba(248, 248, 248, .88);
            border-right: 1px solid var(--win-border);
            backdrop-filter: blur(28px) saturate(125%);
            -webkit-backdrop-filter: blur(28px) saturate(125%);
        }
        section[data-testid="stSidebar"] > div {
            width: var(--sidebar-width) !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarContent"] {
            padding: var(--layout-top) 16px 24px;
        }
        section[data-testid="stSidebar"] hr {
            margin: 14px 0;
            border-color: var(--win-border);
        }
        .sidebar-brand {
            height: 72px;
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 0 54px 0 6px;
            margin: 0 0 14px;
        }
        .sidebar-brand-copy { min-width: 0; }
        .sidebar-brand-title {
            color: var(--win-text);
            font-size: 15px;
            font-weight: 650;
            letter-spacing: -.015em;
            line-height: 1.2;
        }
        .sidebar-brand-subtitle {
            margin-top: 3px;
            color: var(--win-text-secondary);
            font-size: 12px;
            line-height: 1.25;
        }
        .fluent-grid-icon {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 2px;
            width: 32px;
            height: 32px;
            padding: 7px;
            border: 1px solid rgba(0, 103, 192, .18);
            border-radius: 8px;
            background: linear-gradient(145deg, #0a73c9, #005fb8);
            box-shadow: 0 2px 7px rgba(0, 95, 184, .18);
            flex: 0 0 auto;
        }
        .fluent-grid-icon span {
            display: block;
            border-radius: 1px;
            background: rgba(255,255,255,.96);
        }
        .fluent-grid-icon-small { width: 32px; height: 32px; }
        .sidebar-section-title {
            margin: 15px 2px 7px;
            color: var(--win-text-secondary);
            font-size: 11px;
            font-weight: 650;
            letter-spacing: .035em;
            text-transform: uppercase;
        }
        .sidebar-section-help {
            margin: -2px 2px 9px;
            color: var(--win-text-tertiary);
            font-size: 11px;
            line-height: 1.35;
        }
        .sidebar-section-gap { height: 2px; }
        .data-source-card {
            display: flex;
            align-items: center;
            gap: 8px;
            min-height: 34px;
            margin: 8px 0 1px;
            padding: 7px 9px;
            overflow: hidden;
            color: #174b27;
            font-size: 11px;
            line-height: 1.25;
            background: rgba(223, 246, 227, .76);
            border: 1px solid rgba(16, 124, 16, .16);
            border-radius: var(--win-radius-md);
        }
        .data-source-card::before {
            content: "\E73E";
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            color: #107c10;
            font-size: 13px;
            flex: 0 0 auto;
        }
        .data-source-card span {
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }
        .data-source-card-empty {
            color: var(--win-text-secondary);
            background: rgba(255,255,255,.62);
            border-color: var(--win-border);
        }
        .data-source-card-empty::before { content: "\E946"; color: var(--win-text-secondary); }

        section[data-testid="stSidebar"] label {
            color: var(--win-text-secondary) !important;
            font-size: 12px !important;
            font-weight: 600 !important;
        }
        section[data-testid="stSidebar"] [data-baseweb="select"] > div,
        section[data-testid="stSidebar"] [data-testid="stNumberInput"] input,
        section[data-testid="stSidebar"] [data-testid="stTextInput"] input {
            min-height: 34px;
            color: var(--win-text);
            background: rgba(255,255,255,.84);
            border-color: var(--win-border-strong);
            border-radius: var(--win-radius-md);
            box-shadow: 0 1px 1px rgba(0,0,0,.025);
            transition: border-color var(--win-fast) ease, box-shadow var(--win-fast) ease, background var(--win-fast) ease;
        }
        section[data-testid="stSidebar"] [data-baseweb="select"] > div:hover,
        section[data-testid="stSidebar"] [data-testid="stNumberInput"] input:hover,
        section[data-testid="stSidebar"] [data-testid="stTextInput"] input:hover {
            background: #fff;
            border-color: rgba(0,0,0,.2);
        }
        section[data-testid="stSidebar"] [data-baseweb="select"] > div:focus-within,
        section[data-testid="stSidebar"] [data-testid="stNumberInput"] input:focus,
        section[data-testid="stSidebar"] [data-testid="stTextInput"] input:focus {
            border-color: var(--win-accent);
            box-shadow: inset 0 -2px 0 var(--win-accent);
            outline: none;
        }
        section[data-testid="stSidebar"] [data-testid="stFileUploader"] section {
            min-height: 78px;
            padding: 10px;
            background: rgba(255,255,255,.72);
            border: 1px dashed rgba(0,0,0,.22);
            border-radius: var(--win-radius-lg);
            box-shadow: none;
            transition: background var(--win-normal) ease, border-color var(--win-normal) ease, transform var(--win-normal) var(--win-ease);
        }
        section[data-testid="stSidebar"] [data-testid="stFileUploader"] section:hover {
            background: #fff;
            border-color: var(--win-accent);
            transform: translateY(-1px);
        }
        section[data-testid="stSidebar"] [data-testid="stFileUploader"] small { font-size: 10px; }

        /* Main header aligned with the sidebar brand */
        .app-header {
            min-height: 72px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 20px;
            padding: 12px 16px;
            margin: 0 0 12px;
            background: rgba(255,255,255,.68);
            border: 1px solid var(--win-border);
            border-radius: var(--win-radius-xl);
            box-shadow: 0 1px 1px rgba(0,0,0,.025);
            backdrop-filter: blur(24px) saturate(125%);
            -webkit-backdrop-filter: blur(24px) saturate(125%);
        }
        .app-header-main { min-width: 0; }
        .app-title-cluster { display: flex; align-items: center; gap: 11px; }
        .app-product-icon { width: 34px; height: 34px; padding: 8px; border-radius: 9px; }
        .app-title-copy { min-width: 0; }
        .app-eyebrow {
            color: var(--win-accent);
            font-size: 10px;
            font-weight: 650;
            letter-spacing: .07em;
            line-height: 1.15;
            text-transform: uppercase;
        }
        .app-title {
            margin-top: 2px;
            color: var(--win-text);
            font-size: 20px;
            font-weight: 650;
            letter-spacing: -.025em;
            line-height: 1.15;
        }
        .app-subtitle {
            margin: 5px 0 0 45px;
            color: var(--win-text-secondary);
            font-size: 12px;
            line-height: 1.35;
        }
        .app-meta {
            display: flex;
            align-items: center;
            justify-content: flex-end;
            flex-wrap: wrap;
            gap: 6px;
            max-width: 55%;
        }
        .meta-chip {
            min-height: 28px;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            max-width: 260px;
            padding: 5px 9px;
            overflow: hidden;
            color: var(--win-text-secondary);
            font-size: 11px;
            line-height: 1;
            white-space: nowrap;
            text-overflow: ellipsis;
            background: rgba(255,255,255,.76);
            border: 1px solid var(--win-border);
            border-radius: 7px;
        }
        .meta-chip::before {
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            color: var(--win-text-tertiary);
            font-size: 12px;
        }
        .meta-chip-date::before { content: "\E787"; }
        .meta-chip-file::before { content: "\E8A5"; }
        .meta-chip-status::before { content: "\E73E"; color: #107c10; }
        .meta-chip-accent { color: #004f88; background: rgba(224, 240, 255, .82); border-color: rgba(0,103,192,.14); }
        .status-dot { display: none; }

        div[data-testid="stTabs"] {
            margin-top: 0;
        }
        div[data-testid="stTabs"] [role="tablist"] {
            position: sticky;
            top: 8px;
            z-index: 80;
            display: flex;
            align-items: center;
            justify-content: flex-start;
            gap: 4px !important;
            min-height: 46px;
            margin: 0 0 14px;
            padding: 5px;
            overflow-x: auto;
            overflow-y: hidden;
            scrollbar-width: none;
            background: rgba(248, 248, 248, .90);
            border: 1px solid var(--win-border);
            border-radius: 10px;
            box-shadow: 0 3px 14px rgba(0, 0, 0, .055);
            backdrop-filter: blur(24px) saturate(135%);
            -webkit-backdrop-filter: blur(24px) saturate(135%);
        }
        div[data-testid="stTabs"] [role="tablist"]::-webkit-scrollbar {
            display: none;
        }
        div[data-testid="stTabs"] button[role="tab"] {
            position: relative;
            flex: 0 0 auto !important;
            width: auto !important;
            min-width: max-content !important;
            min-height: 36px;
            display: inline-flex !important;
            align-items: center;
            justify-content: center;
            column-gap: 8px !important;
            margin: 0 !important;
            padding: 0 15px !important;
            color: var(--win-text-secondary);
            font-family: inherit;
            font-size: 12px;
            font-weight: 600;
            letter-spacing: 0;
            white-space: nowrap !important;
            background: transparent;
            border: 1px solid transparent !important;
            border-radius: 7px;
            box-shadow: none;
            transition:
                color var(--win-fast) ease,
                background var(--win-fast) ease,
                border-color var(--win-fast) ease,
                box-shadow var(--win-normal) var(--win-ease),
                transform 80ms ease;
        }
        div[data-testid="stTabs"] button[role="tab"] p,
        div[data-testid="stTabs"] button[role="tab"] span {
            margin: 0 !important;
            padding: 0 !important;
            color: inherit !important;
            font-size: inherit !important;
            font-weight: inherit !important;
            line-height: 1 !important;
            white-space: nowrap !important;
        }
        div[data-testid="stTabs"] button[role="tab"]::before {
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            color: currentColor;
            font-size: 13px;
            font-weight: 400;
            opacity: .88;
            transition:
                color var(--win-fast) ease,
                transform var(--win-normal) var(--win-ease),
                opacity var(--win-fast) ease;
        }
        div[data-testid="stTabs"] button[role="tab"]:nth-of-type(1)::before { content: "\E80F"; }
        div[data-testid="stTabs"] button[role="tab"]:nth-of-type(2)::before { content: "\E77B"; }
        div[data-testid="stTabs"] button[role="tab"]:nth-of-type(3)::before { content: "\E721"; }
        div[data-testid="stTabs"] button[role="tab"]:nth-of-type(4)::before { content: "\E9D5"; }
        div[data-testid="stTabs"] button[role="tab"]:nth-of-type(5)::before { content: "\E9D9"; }
        div[data-testid="stTabs"] button[role="tab"]:nth-of-type(6)::before { content: "\E897"; }
        div[data-testid="stTabs"] button[role="tab"]:hover:not([aria-selected="true"]) {
            color: var(--win-text);
            background: rgba(0, 0, 0, .045);
            border-color: rgba(0, 0, 0, .025) !important;
        }
        div[data-testid="stTabs"] button[role="tab"]:hover:not([aria-selected="true"])::before {
            color: var(--win-accent);
            opacity: 1;
            transform: translateY(-1px);
        }
        div[data-testid="stTabs"] button[role="tab"]:active {
            transform: scale(.985);
            background: rgba(0, 0, 0, .065);
        }
        div[data-testid="stTabs"] button[role="tab"]:focus-visible {
            outline: 2px solid rgba(0, 103, 192, .62);
            outline-offset: 1px;
        }
        div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
            color: var(--win-text) !important;
            background: rgba(255, 255, 255, .96) !important;
            border-color: rgba(0, 0, 0, .065) !important;
            box-shadow: 0 1px 3px rgba(0, 0, 0, .09);
            animation: fluentTabSelect var(--win-normal) var(--win-ease) both;
        }
        div[data-testid="stTabs"] button[role="tab"][aria-selected="true"]::before {
            color: var(--win-accent);
            opacity: 1;
        }
        div[data-testid="stTabs"] button[role="tab"][aria-selected="true"]::after {
            content: "";
            position: absolute;
            left: 50%;
            bottom: 1px;
            width: 18px;
            height: 2px;
            border-radius: 999px;
            background: var(--win-accent);
            transform: translateX(-50%);
            animation: fluentTabIndicator var(--win-normal) var(--win-ease) both;
        }
        @keyframes fluentTabSelect {
            from { opacity: .78; transform: translateY(1px) scale(.99); }
            to { opacity: 1; transform: translateY(0) scale(1); }
        }
        @keyframes fluentTabIndicator {
            from { width: 6px; opacity: .35; }
            to { width: 18px; opacity: 1; }
        }
        div[data-testid="stTabs"] [data-baseweb="tab-highlight"],
        div[data-testid="stTabs"] [data-baseweb="tab-border"] {
            display: none !important;
            width: 0 !important;
            height: 0 !important;
            border: 0 !important;
        }
        div[data-testid="stTabs"] [role="tabpanel"] {
            position: relative;
            z-index: 1;
            padding-top: 2px;
            animation: fluentContentReveal 200ms var(--win-ease) both;
        }
        @keyframes fluentContentReveal {
            from { opacity: 0; transform: translateY(3px); }
            to { opacity: 1; transform: translateY(0); }
        }

        /* Content hierarchy */
        .tab-page-heading {
            margin: 1px 0 13px;
        }
        .tab-page-title, .section-title {
            color: var(--win-text);
            font-size: 17px;
            font-weight: 650;
            letter-spacing: -.018em;
            line-height: 1.22;
        }
        .tab-page-subtitle, .section-subtitle {
            margin-top: 3px;
            color: var(--win-text-secondary);
            font-size: 12px;
            line-height: 1.38;
        }
        .section-title { font-size: 14px; }
        .section-subtitle { margin-bottom: 8px; }
        .compact-heading { margin-top: 0; }
        .section-divider {
            height: 1px;
            margin: 16px 0 13px;
            background: var(--win-border);
        }
        .section-block { height: 14px; }
        .kpi-row-gap { height: 8px; }

        /* KPI cards */
        .kpi-card {
            position: relative;
            min-height: 104px;
            padding: 13px 14px 12px;
            overflow: hidden;
            background: var(--win-surface);
            border: 1px solid var(--win-border);
            border-radius: var(--win-radius-lg);
            box-shadow: var(--win-shadow-card);
            transition: transform var(--win-normal) var(--win-ease), box-shadow var(--win-normal) var(--win-ease), border-color var(--win-normal) ease, background var(--win-normal) ease;
        }
        .kpi-card:hover {
            transform: translateY(-2px);
            background: #fff;
            border-color: rgba(0,0,0,.10);
            box-shadow: 0 2px 4px rgba(0,0,0,.045), 0 9px 24px rgba(0,0,0,.075);
        }
        .kpi-accent {
            position: absolute;
            inset: 0 auto 0 0;
            width: 3px;
            background: var(--win-accent);
        }
        .kpi-topline { display: flex; align-items: center; justify-content: space-between; gap: 10px; }
        .kpi-label {
            color: var(--win-text-secondary);
            font-size: 11px;
            font-weight: 600;
            line-height: 1.2;
        }
        .kpi-icon {
            width: 24px;
            height: 24px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            color: var(--win-accent);
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            font-size: 13px;
            background: var(--win-accent-soft);
            border-radius: 6px;
            flex: 0 0 auto;
        }
        .kpi-value {
            margin-top: 7px;
            color: var(--win-text);
            font-size: 24px;
            font-weight: 650;
            letter-spacing: -.035em;
            line-height: 1;
        }
        .kpi-help {
            margin-top: 7px;
            overflow: hidden;
            color: var(--win-text-tertiary);
            font-size: 10.5px;
            line-height: 1.25;
            text-overflow: ellipsis;
            white-space: nowrap;
        }
        .kpi-tone-danger .kpi-accent { background: #c42b1c; }
        .kpi-tone-danger .kpi-icon { color: #c42b1c; background: rgba(196,43,28,.09); }
        .kpi-tone-warning .kpi-accent { background: #d83b01; }
        .kpi-tone-warning .kpi-icon { color: #b83b00; background: rgba(216,59,1,.09); }
        .kpi-tone-watch .kpi-accent { background: #9d7b00; }
        .kpi-tone-watch .kpi-icon { color: #806000; background: rgba(157,123,0,.10); }
        .kpi-tone-success .kpi-accent { background: #107c10; }
        .kpi-tone-success .kpi-icon { color: #107c10; background: rgba(16,124,16,.09); }
        .kpi-tone-neutral .kpi-accent { background: #6b6b6b; }
        .kpi-tone-neutral .kpi-icon { color: #5c5c5c; background: rgba(0,0,0,.055); }

        /* Selected SKU */
        .selected-sku-card {
            position: relative;
            margin: 0 0 12px;
            padding: 15px 17px;
            overflow: hidden;
            background: rgba(255,255,255,.82);
            border: 1px solid var(--win-border);
            border-radius: var(--win-radius-lg);
            box-shadow: var(--win-shadow-card);
        }
        .selected-sku-card::before {
            content: "";
            position: absolute;
            inset: 0 auto 0 0;
            width: 3px;
            background: var(--win-accent);
        }
        .selected-sku-label {
            color: var(--win-text-secondary);
            font-size: 10px;
            font-weight: 650;
            letter-spacing: .04em;
            text-transform: uppercase;
        }
        .selected-sku-value {
            margin-top: 4px;
            color: var(--win-text);
            font-size: 25px;
            font-weight: 650;
            letter-spacing: -.035em;
            line-height: 1.05;
        }
        .selected-sku-description {
            margin-top: 5px;
            color: var(--win-text-secondary);
            font-size: 12px;
            line-height: 1.35;
        }

        /* Streamlit controls */
        .stButton > button, .stDownloadButton > button {
            min-height: 34px;
            padding: 6px 12px;
            color: var(--win-text);
            font-family: inherit;
            font-size: 12px;
            font-weight: 600;
            background: rgba(255,255,255,.88);
            border: 1px solid var(--win-border-strong);
            border-radius: var(--win-radius-md);
            box-shadow: 0 1px 2px rgba(0,0,0,.035);
            transition: background var(--win-fast) ease, border-color var(--win-fast) ease, box-shadow var(--win-fast) ease, transform 80ms ease;
        }
        .stButton > button:hover, .stDownloadButton > button:hover {
            color: var(--win-text);
            background: #fff;
            border-color: rgba(0,0,0,.20);
            box-shadow: 0 2px 7px rgba(0,0,0,.08);
        }
        .stButton > button:active, .stDownloadButton > button:active {
            transform: scale(.985);
            background: #f6f6f6;
        }
        .stButton > button:focus-visible, .stDownloadButton > button:focus-visible {
            outline: 2px solid rgba(0,103,192,.58);
            outline-offset: 2px;
        }
        .stDownloadButton > button::before {
            content: "\E896";
            margin-right: 7px;
            color: var(--win-accent);
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            font-size: 12px;
            font-weight: 400;
        }
        section[data-testid="stSidebar"] .stButton > button {
            color: var(--win-text-secondary);
            background: transparent;
            border-color: transparent;
            box-shadow: none;
        }
        section[data-testid="stSidebar"] .stButton { margin-top: 5px; }
        section[data-testid="stSidebar"] .stButton > button:hover {
            color: var(--win-text);
            background: rgba(0,0,0,.045);
        }
        div[data-testid="stTextArea"] textarea,
        div[data-testid="stTextInput"] input,
        div[data-testid="stDateInput"] input,
        div[data-testid="stNumberInput"] input,
        div[data-baseweb="select"] > div {
            color: var(--win-text);
            background: rgba(255,255,255,.88);
            border-color: var(--win-border-strong);
            border-radius: var(--win-radius-md);
            transition: border-color var(--win-fast) ease, box-shadow var(--win-fast) ease, background var(--win-fast) ease;
        }
        div[data-testid="stTextArea"] textarea:hover,
        div[data-testid="stTextInput"] input:hover,
        div[data-testid="stDateInput"] input:hover,
        div[data-testid="stNumberInput"] input:hover,
        div[data-baseweb="select"] > div:hover { background: #fff; border-color: rgba(0,0,0,.2); }
        div[data-testid="stTextArea"] textarea:focus,
        div[data-testid="stTextInput"] input:focus,
        div[data-testid="stDateInput"] input:focus,
        div[data-testid="stNumberInput"] input:focus,
        div[data-baseweb="select"] > div:focus-within {
            border-color: var(--win-accent);
            box-shadow: inset 0 -2px 0 var(--win-accent);
            outline: none;
        }
        div[data-testid="stTextArea"] textarea { min-height: 104px !important; }
        div[data-testid="stRadio"] label { font-size: 12px !important; }
        div[data-testid="stCheckbox"] label { font-size: 12px !important; }

        /* Tables */
        div[data-testid="stDataFrame"] {
            margin-top: 5px;
            overflow: hidden;
            background: rgba(255,255,255,.84);
            border: 1px solid var(--win-border);
            border-radius: var(--win-radius-lg);
            box-shadow: 0 1px 2px rgba(0,0,0,.025);
            transition: border-color var(--win-normal) ease, box-shadow var(--win-normal) ease;
        }
        div[data-testid="stDataFrame"]:hover {
            border-color: rgba(0,0,0,.10);
            box-shadow: 0 3px 14px rgba(0,0,0,.05);
        }
        div[data-testid="stCaptionContainer"] {
            margin: 3px 0 2px;
            color: var(--win-text-tertiary);
            font-size: 10.5px;
        }

        /* Expanders and alerts */
        div[data-testid="stExpander"] {
            margin-bottom: 8px;
            overflow: hidden;
            background: rgba(255,255,255,.72);
            border: 1px solid var(--win-border);
            border-radius: var(--win-radius-lg);
            box-shadow: none;
            transition: background var(--win-normal) ease, border-color var(--win-normal) ease, box-shadow var(--win-normal) ease;
        }
        div[data-testid="stExpander"]:hover {
            background: rgba(255,255,255,.90);
            border-color: rgba(0,0,0,.10);
            box-shadow: 0 2px 10px rgba(0,0,0,.045);
        }
        div[data-testid="stExpander"] details[open] { animation: contentReveal .18s var(--win-ease); }
        div[data-testid="stAlert"] {
            border-radius: var(--win-radius-lg);
            border-width: 1px;
            box-shadow: none;
        }

        /* Search/status components */
        .tx-filter-shell, .lookup-hero, .stock-input-example, .loading-stage-card, .ready-stage-card, .error-stage-card, .empty-state {
            background: rgba(255,255,255,.78);
            border: 1px solid var(--win-border);
            border-radius: var(--win-radius-lg);
            box-shadow: var(--win-shadow-card);
        }
        .tx-filter-shell { padding: 14px; margin: 0 0 11px; }
        .tx-filter-card { min-height: 0; padding: 11px 12px; background: rgba(255,255,255,.66); border: 1px solid var(--win-border); border-radius: 9px; box-shadow: none; }
        .tx-filter-title { color: var(--win-text); font-size: 15px; font-weight: 650; }
        .tx-filter-subtitle, .tx-filter-card-subtitle { color: var(--win-text-secondary); font-size: 11px; }
        .tx-filter-card-title { color: var(--win-text-secondary); font-size: 10px; font-weight: 650; letter-spacing: .04em; }
        .tx-status-grid { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 8px; margin: 10px 0 8px; }
        .tx-status-card {
            min-height: 70px;
            padding: 10px 11px;
            background: rgba(255,255,255,.82);
            border: 1px solid var(--win-border);
            border-radius: 9px;
            box-shadow: none;
            transition: transform var(--win-normal) var(--win-ease), background var(--win-normal) ease;
        }
        .tx-status-card:hover { transform: translateY(-1px); background: #fff; }
        .tx-status-label { color: var(--win-text-tertiary); font-size: 9.5px; font-weight: 650; letter-spacing: .04em; }
        .tx-status-value { margin-top: 4px; color: var(--win-text); font-size: 18px; font-weight: 650; }
        .tx-status-help { margin-top: 4px; color: var(--win-text-tertiary); font-size: 10px; }
        .tx-result-box { margin: 8px 0 6px; padding: 10px 11px; background: rgba(255,255,255,.72); border: 1px solid var(--win-border); border-radius: 9px; }
        .tx-result-box-ok { background: rgba(223,246,227,.64); border-color: rgba(16,124,16,.15); }
        .tx-result-box-missing { background: rgba(253,231,233,.72); border-color: rgba(196,43,28,.15); }
        .tx-result-title { color: var(--win-text); font-size: 11px; font-weight: 650; }
        .tx-pill-wrap { display: flex; flex-wrap: wrap; gap: 5px; margin-top: 6px; }
        .tx-example-pill, .tx-pill, .tx-pill-ok, .tx-pill-missing, .tx-pill-muted {
            display: inline-flex;
            align-items: center;
            min-height: 23px;
            padding: 3px 7px;
            color: var(--win-text-secondary);
            font-size: 10.5px;
            font-weight: 600;
            background: rgba(0,0,0,.04);
            border: 1px solid var(--win-border);
            border-radius: 999px;
        }
        .tx-pill-ok { color: #0f6c0f; background: rgba(223,246,227,.8); border-color: rgba(16,124,16,.15); }
        .tx-pill-missing { color: #a4262c; background: rgba(253,231,233,.9); border-color: rgba(196,43,28,.15); }
        .tx-pill-muted { color: var(--win-text-secondary); background: rgba(0,0,0,.035); }
        .stock-input-example { min-height: 0; padding: 11px 12px; box-shadow: none; }
        .stock-input-code { margin-top: 7px; padding: 8px 9px; color: var(--win-text-secondary); font-size: 10.5px; background: rgba(0,0,0,.035); border: 1px solid var(--win-border); border-radius: 7px; }
        .stock-do-heading { color: var(--win-text); font-size: 13px; font-weight: 650; }
        .stock-do-subtitle { color: var(--win-text-secondary); font-size: 11px; }
        .empty-state { max-width: 620px; margin: 64px auto; padding: 28px; text-align: center; }
        .empty-state::before {
            content: "\E898";
            display: block;
            margin-bottom: 12px;
            color: var(--win-accent);
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            font-size: 30px;
        }
        .empty-state-title { color: var(--win-text); font-size: 17px; font-weight: 650; }
        .empty-state-text { margin-top: 6px; color: var(--win-text-secondary); font-size: 12px; line-height: 1.45; }
        .loading-stage-card, .ready-stage-card, .error-stage-card {
            position: relative;
            min-height: 74px;
            padding: 14px 15px;
            overflow: hidden;
            background: rgba(255,255,255,.88);
            border: 1px solid var(--win-border);
            border-radius: var(--win-radius-lg);
            box-shadow: 0 1px 2px rgba(0,0,0,.035), 0 8px 24px rgba(0,0,0,.055);
            backdrop-filter: blur(20px) saturate(125%);
            -webkit-backdrop-filter: blur(20px) saturate(125%);
            animation: uploadCardIn .26s var(--win-ease) both;
        }
        .loading-row, .ready-row, .error-row {
            position: relative;
            z-index: 1;
            display: flex;
            align-items: center;
            gap: 12px;
        }
        .loader-ring {
            width: 28px;
            height: 28px;
            flex: 0 0 auto;
            border: 2.5px solid rgba(0,103,192,.14);
            border-top-color: var(--win-accent);
            border-radius: 50%;
            animation: uploadSpin .9s linear infinite;
        }
        .ready-check, .error-mark {
            width: 28px;
            height: 28px;
            display: flex;
            align-items: center;
            justify-content: center;
            flex: 0 0 auto;
            color: #fff;
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets", sans-serif;
            font-size: 13px;
            font-weight: 600;
            border-radius: 50%;
        }
        .ready-check {
            background: #107c10;
            box-shadow: 0 4px 12px rgba(16,124,16,.18);
            animation: readyPop .3s var(--win-ease) both;
        }
        .error-mark {
            background: #c42b1c;
            box-shadow: 0 4px 12px rgba(196,43,28,.16);
        }
        .animated-progress {
            position: relative;
            z-index: 1;
            height: 3px;
            margin-top: 13px;
            overflow: hidden;
            background: rgba(0,103,192,.10);
            border-radius: 999px;
        }
        .animated-progress::before {
            content: "";
            position: absolute;
            inset: 0;
            width: 34%;
            border-radius: inherit;
            background: linear-gradient(90deg, transparent, var(--win-accent), transparent);
            animation: uploadProgress 1.35s var(--win-ease) infinite;
        }
        .ready-stage-card {
            border-color: rgba(16,124,16,.16);
            background: rgba(247,253,248,.94);
            animation: readyCardIn .3s var(--win-ease) both, readyCardOut .34s ease 1.35s forwards;
        }
        .error-stage-card {
            border-color: rgba(196,43,28,.16);
            background: rgba(255,248,248,.96);
        }
        .stage-copy { min-width: 0; }
        .stage-title {
            color: var(--win-text);
            font-size: 12px;
            font-weight: 650;
            line-height: 1.25;
        }
        .stage-subtitle {
            margin-top: 3px;
            color: var(--win-text-secondary);
            font-size: 11px;
            line-height: 1.38;
        }
        .stage-meta {
            margin-top: 4px;
            color: var(--win-text-tertiary);
            font-size: 10px;
            line-height: 1.3;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        @keyframes uploadSpin {
            to { transform: rotate(360deg); }
        }
        @keyframes uploadProgress {
            from { transform: translateX(-130%); }
            to { transform: translateX(390%); }
        }
        @keyframes uploadCardIn {
            from { opacity: 0; transform: translateY(4px); }
            to { opacity: 1; transform: translateY(0); }
        }
        @keyframes readyCardIn {
            from { opacity: 0; transform: translateY(3px); }
            to { opacity: 1; transform: translateY(0); }
        }
        @keyframes readyCardOut {
            0% { opacity: 1; max-height: 120px; margin-bottom: 0; padding-top: 14px; padding-bottom: 14px; }
            70% { opacity: 0; max-height: 120px; margin-bottom: 0; padding-top: 14px; padding-bottom: 14px; }
            100% { opacity: 0; max-height: 0; min-height: 0; margin: 0; padding-top: 0; padding-bottom: 0; border-width: 0; }
        }
        @keyframes readyPop {
            0% { opacity: 0; transform: scale(.82); }
            65% { opacity: 1; transform: scale(1.06); }
            100% { opacity: 1; transform: scale(1); }
        }
        .tx-filter-top { display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; margin-bottom: 9px; }
        .tx-example-row { display: flex; flex-wrap: wrap; gap: 5px; margin: 5px 0 8px; }
        .stButton > button[kind="primary"], .stDownloadButton > button[kind="primary"] {
            color: #fff;
            background: var(--win-accent);
            border-color: var(--win-accent);
            box-shadow: 0 1px 2px rgba(0,0,0,.08);
        }
        .stButton > button[kind="primary"]:hover, .stDownloadButton > button[kind="primary"]:hover {
            color: #fff;
            background: var(--win-accent-hover);
            border-color: var(--win-accent-hover);
        }
        .stButton > button[kind="primary"]:active, .stDownloadButton > button[kind="primary"]:active {
            background: var(--win-accent-pressed);
            border-color: var(--win-accent-pressed);
        }
        .stDownloadButton > button[kind="primary"]::before { color: #fff; }
        .loader-ring { border-color: rgba(0,103,192,.14); border-top-color: var(--win-accent); }
        .animated-progress { height: 4px; background: rgba(0,103,192,.10); }
        .animated-progress::before { background: linear-gradient(90deg, transparent, var(--win-accent), transparent); }
        .stage-title { color: var(--win-text); font-size: 12px; font-weight: 650; }
        .stage-subtitle { color: var(--win-text-secondary); font-size: 11px; }

        /* Responsive */
        @media (max-width: 1180px) {
            .app-header { align-items: flex-start; flex-direction: column; gap: 9px; }
            .app-meta { justify-content: flex-start; max-width: 100%; margin-left: 45px; }
            .tx-status-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
        }
        @media (max-width: 820px) {
            :root { --layout-x: 14px; --layout-top: 6px; }
            .app-header { padding: 12px; border-radius: 12px; }
            .app-subtitle, .app-meta { margin-left: 0; }
            .app-meta { display: grid; grid-template-columns: 1fr; width: 100%; }
            .meta-chip { width: 100%; max-width: none; }
            div[data-testid="stTabs"] [role="tablist"] { top: 6px; }
            .kpi-card { min-height: 96px; }
        }
        @media (max-width: 520px) {
            .main .block-container { padding-left: 9px; padding-right: 9px; }
            .app-product-icon { display: none; }
            .app-title { font-size: 18px; }
            .app-subtitle { font-size: 11px; }
            div[data-testid="stTabs"] [role="tablist"] { gap: 3px !important; padding-left: 7px; padding-right: 7px; }
            div[data-testid="stTabs"] button[role="tab"] { padding: 0 13px !important; font-size: 11.5px; }
            .tx-status-grid { grid-template-columns: 1fr; }
        }


        .st-key-main_navigation {
            position: sticky;
            top: 8px;
            z-index: 80;
            margin: 0 0 16px;
            padding: 4px;
            overflow-x: auto;
            scrollbar-width: none;
            background: rgba(248,248,248,.90);
            border: 1px solid var(--win-border);
            border-radius: 10px;
            box-shadow: 0 3px 14px rgba(0,0,0,.055);
            backdrop-filter: blur(24px) saturate(135%);
            -webkit-backdrop-filter: blur(24px) saturate(135%);
        }
        .st-key-main_navigation::-webkit-scrollbar { display: none; }
        .st-key-main_navigation [data-testid="stSegmentedControl"] { min-width: max-content; }
        .st-key-main_navigation [data-testid="stSegmentedControl"] > div,
        .st-key-main_navigation [role="radiogroup"] {
            display: flex;
            align-items: center;
            justify-content: flex-start;
            gap: 5px;
            min-width: max-content;
            padding: 0;
            background: transparent;
            border: 0;
        }
        .st-key-main_navigation button,
        .st-key-main_navigation [role="radio"] {
            position: relative;
            min-width: max-content;
            min-height: 36px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 7px;
            padding: 0 15px !important;
            color: var(--win-text-secondary) !important;
            font-family: inherit !important;
            font-size: 12px !important;
            font-weight: 600 !important;
            white-space: nowrap;
            background: transparent !important;
            border: 1px solid transparent !important;
            border-radius: 7px !important;
            box-shadow: none !important;
            transition: color var(--win-fast) ease, background var(--win-fast) ease, border-color var(--win-fast) ease, box-shadow var(--win-normal) var(--win-ease), transform 80ms ease !important;
        }
        .st-key-main_navigation button::before,
        .st-key-main_navigation [role="radio"]::before {
            font-family: "Segoe Fluent Icons", "Segoe MDL2 Assets";
            color: currentColor;
            font-size: 13px;
            font-weight: 400;
            opacity: .88;
        }
        .st-key-main_navigation button:nth-of-type(1)::before,
        .st-key-main_navigation [role="radio"]:nth-of-type(1)::before { content: "\E80F"; }
        .st-key-main_navigation button:nth-of-type(2)::before,
        .st-key-main_navigation [role="radio"]:nth-of-type(2)::before { content: "\E77B"; }
        .st-key-main_navigation button:nth-of-type(3)::before,
        .st-key-main_navigation [role="radio"]:nth-of-type(3)::before { content: "\E721"; }
        .st-key-main_navigation button:nth-of-type(4)::before,
        .st-key-main_navigation [role="radio"]:nth-of-type(4)::before { content: "\E9D5"; }
        .st-key-main_navigation button:nth-of-type(5)::before,
        .st-key-main_navigation [role="radio"]:nth-of-type(5)::before { content: "\E9D9"; }
        .st-key-main_navigation button:nth-of-type(6)::before,
        .st-key-main_navigation [role="radio"]:nth-of-type(6)::before { content: "\E897"; }
        .st-key-main_navigation button:hover,
        .st-key-main_navigation [role="radio"]:hover {
            color: var(--win-text) !important;
            background: rgba(0,0,0,.045) !important;
        }
        .st-key-main_navigation button:active,
        .st-key-main_navigation [role="radio"]:active { transform: scale(.985); }
        .st-key-main_navigation button:focus-visible,
        .st-key-main_navigation [role="radio"]:focus-visible {
            outline: 2px solid rgba(0,103,192,.58) !important;
            outline-offset: 1px !important;
        }
        .st-key-main_navigation button[aria-pressed="true"],
        .st-key-main_navigation button[data-selected="true"],
        .st-key-main_navigation [role="radio"][aria-checked="true"] {
            color: var(--win-text) !important;
            background: #fff !important;
            border-color: rgba(0,0,0,.06) !important;
            box-shadow: 0 1px 3px rgba(0,0,0,.09) !important;
        }
        .st-key-main_navigation button[aria-pressed="true"]::before,
        .st-key-main_navigation button[data-selected="true"]::before,
        .st-key-main_navigation [role="radio"][aria-checked="true"]::before { color: var(--win-accent); }
        .st-key-main_navigation button[aria-pressed="true"]::after,
        .st-key-main_navigation button[data-selected="true"]::after,
        .st-key-main_navigation [role="radio"][aria-checked="true"]::after {
            content: "";
            position: absolute;
            left: 50%;
            bottom: 1px;
            width: 18px;
            height: 2px;
            border-radius: 999px;
            background: var(--win-accent);
            transform: translateX(-50%);
            animation: tabIndicator .18s var(--win-ease) both;
        }
        .st-key-main_navigation p { margin: 0 !important; color: inherit !important; font: inherit !important; }
        @media (max-width: 820px) {
            .st-key-main_navigation { top: 6px; }
            .st-key-main_navigation button,
            .st-key-main_navigation [role="radio"] { padding: 0 12px !important; }
        }

        @media (prefers-reduced-motion: reduce) {
            html { scroll-behavior: auto; }
            *, *::before, *::after {
                animation-duration: .01ms !important;
                animation-iteration-count: 1 !important;
                transition-duration: .01ms !important;
                scroll-behavior: auto !important;
            }
        }
    </style>
    """,
    unsafe_allow_html=True,
)

FORMAT_CONFIGS = {
    "Newark": {
        "title": "Inventory Shortage",
        "sidebar_title": "Inventory Dashboard",
        "caption": "Upload an Item Activity Report Excel file to generate the shortage dashboard.",
        "upload_label": "Drop Item Activity Report here",
        "placeholder": "Search SKU...",
        "help": "Select the matching report format before uploading the Excel file.",
        "cols": {
            "sku": 0,
            "description": 2,
            "activity_date": 7,
            "trans_no": 9,
            "ref_no": 10,
            "qty_in": 12,
            "qty_out": 14,
            "balance": 19,
            "ctn_balance": 20,
        },
        "total_rule": "ref_total",
        "total_source": "Ref # = Total",
        "wrong_format_warning": "Wrong file format for Newark. Switch to Carson or upload the Newark report.",
    },
    "Carson": {
        "title": "Inventory Shortage",
        "sidebar_title": "Inventory Dashboard",
        "caption": "Upload an Item Activity Report Excel file to generate the shortage dashboard.",
        "upload_label": "Drop Item Activity Report here",
        "placeholder": "Search SKU...",
        "help": "Select the matching report format before uploading the Excel file.",
        "cols": {
            "sku": 1,
            "description": 3,
            "activity_date": 5,
            "trans_no": 6,
            "ref_no": 7,
            "qty_in": 8,
            "qty_out": 10,
            "balance": 12,
            "ctn_balance": 13,
        },
        "total_rule": "sku_totals",
        "total_source": "Column B = Totals:",
        "wrong_format_warning": "Wrong file format for Carson. Switch to Newark or upload the Carson report.",
    },
}


def clean_text(value) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    return str(value).replace("\u00a0", " ").replace("\u200b", "").strip()


def get_cell(row, col_idx):
    if col_idx is None:
        return None
    if len(row) <= col_idx:
        return None
    return row.iloc[col_idx] if hasattr(row, "iloc") else row[col_idx]


def first_qty_number(value, default=0.0) -> float:
    text = clean_text(value)
    if not text:
        return default
    text = text.replace(",", "")
    before_slash = text.split("/")[0]
    match = re.search(r"-?\d+(?:\.\d+)?", before_slash)
    return float(match.group()) if match else default


def parse_excel_or_text_date(value):
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return pd.NaT

    if isinstance(value, (datetime, pd.Timestamp)):
        return pd.to_datetime(value).normalize()

    if isinstance(value, date):
        return pd.to_datetime(value).normalize()

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value > 30000:
            return pd.to_datetime(value, unit="D", origin="1899-12-30").normalize()
        return pd.NaT

    text = clean_text(value)
    if not text:
        return pd.NaT

    text = re.sub(r"\s*\([^)]*\)", "", text).strip()
    parsed = pd.to_datetime(text, errors="coerce")
    return parsed.normalize() if not pd.isna(parsed) else pd.NaT


class WrongFileFormatError(ValueError):
    pass


PERSISTENT_UPLOAD_DIR = Path.home() / ".inventory_dashboard_uploads"
PERSISTENT_APP_STATE_PATH = PERSISTENT_UPLOAD_DIR / "app_state.json"
PERSISTENT_APP_STATE_CACHE_KEY = "_inventory_dashboard_app_state_raw"
PERSISTENT_APP_STATE_RESTORED_KEY = "_inventory_dashboard_app_state_restored"


def encode_persistent_value(value):
    if isinstance(value, pd.DataFrame):
        frame = value.copy().astype(object)
        frame = frame.where(pd.notna(frame), None)
        return {
            "__type__": "dataframe",
            "columns": [str(column) for column in frame.columns],
            "records": [
                {str(key): encode_persistent_value(item) for key, item in record.items()}
                for record in frame.to_dict(orient="records")
            ],
        }
    if isinstance(value, pd.Timestamp):
        return {"__type__": "datetime", "value": value.isoformat()}
    if isinstance(value, datetime):
        return {"__type__": "datetime", "value": value.isoformat()}
    if isinstance(value, date):
        return {"__type__": "date", "value": value.isoformat()}
    if isinstance(value, tuple):
        return {"__type__": "tuple", "items": [encode_persistent_value(item) for item in value]}
    if isinstance(value, set):
        return {"__type__": "set", "items": [encode_persistent_value(item) for item in value]}
    if isinstance(value, list):
        return [encode_persistent_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): encode_persistent_value(item) for key, item in value.items()}
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return None if np.isnan(value) else float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    if value is pd.NA:
        return None
    return value


def decode_persistent_value(value):
    if isinstance(value, list):
        return [decode_persistent_value(item) for item in value]
    if not isinstance(value, dict):
        return value
    value_type = value.get("__type__")
    if value_type == "dataframe":
        records = [
            {key: decode_persistent_value(item) for key, item in record.items()}
            for record in value.get("records", [])
        ]
        return pd.DataFrame(records, columns=value.get("columns", []))
    if value_type == "datetime":
        return pd.to_datetime(value.get("value"), errors="coerce")
    if value_type == "date":
        parsed = pd.to_datetime(value.get("value"), errors="coerce")
        return None if pd.isna(parsed) else parsed.date()
    if value_type == "tuple":
        return tuple(decode_persistent_value(item) for item in value.get("items", []))
    if value_type == "set":
        return set(decode_persistent_value(item) for item in value.get("items", []))
    return {key: decode_persistent_value(item) for key, item in value.items()}


def load_persistent_app_state() -> dict:
    try:
        if not PERSISTENT_APP_STATE_PATH.exists():
            st.session_state[PERSISTENT_APP_STATE_CACHE_KEY] = {}
            return {}
        raw_state = json.loads(PERSISTENT_APP_STATE_PATH.read_text(encoding="utf-8"))
        st.session_state[PERSISTENT_APP_STATE_CACHE_KEY] = raw_state
        return {key: decode_persistent_value(value) for key, value in raw_state.items()}
    except Exception:
        st.session_state[PERSISTENT_APP_STATE_CACHE_KEY] = {}
        return {}


def update_persistent_app_state(values=None, remove_keys=None) -> None:
    try:
        PERSISTENT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        cached_state = st.session_state.get(PERSISTENT_APP_STATE_CACHE_KEY)
        if isinstance(cached_state, dict):
            raw_state = dict(cached_state)
        elif PERSISTENT_APP_STATE_PATH.exists():
            loaded = json.loads(PERSISTENT_APP_STATE_PATH.read_text(encoding="utf-8"))
            raw_state = loaded if isinstance(loaded, dict) else {}
        else:
            raw_state = {}
        changed = False
        for key in remove_keys or []:
            if key in raw_state:
                raw_state.pop(key, None)
                changed = True
        for key, value in (values or {}).items():
            encoded_value = encode_persistent_value(value)
            if key not in raw_state or raw_state[key] != encoded_value:
                raw_state[key] = encoded_value
                changed = True
        if not changed:
            st.session_state[PERSISTENT_APP_STATE_CACHE_KEY] = raw_state
            return
        temp_path = PERSISTENT_APP_STATE_PATH.with_suffix(".tmp")
        temp_path.write_text(json.dumps(raw_state, ensure_ascii=False), encoding="utf-8")
        temp_path.replace(PERSISTENT_APP_STATE_PATH)
        st.session_state[PERSISTENT_APP_STATE_CACHE_KEY] = raw_state
    except Exception:
        pass


def restore_persistent_app_state() -> None:
    if st.session_state.get(PERSISTENT_APP_STATE_RESTORED_KEY):
        return
    for key, value in load_persistent_app_state().items():
        if key not in st.session_state:
            st.session_state[key] = value
    st.session_state[PERSISTENT_APP_STATE_RESTORED_KEY] = True


def stable_file_hash(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def safe_format_slug(format_name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_-]+", "_", str(format_name)).strip("_") or "report"


def persistent_upload_paths(format_name: str) -> tuple[Path, Path]:
    slug = safe_format_slug(format_name)
    return (
        PERSISTENT_UPLOAD_DIR / f"{slug}_upload.bin",
        PERSISTENT_UPLOAD_DIR / f"{slug}_upload.json",
    )

def save_persistent_upload(format_name: str, file_name: str, file_bytes: bytes, file_hash: str = "") -> None:
    try:
        PERSISTENT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        data_path, meta_path = persistent_upload_paths(format_name)
        resolved_hash = file_hash or stable_file_hash(file_bytes)
        if data_path.exists() and meta_path.exists() and data_path.stat().st_size == len(file_bytes):
            existing_meta = json.loads(meta_path.read_text(encoding="utf-8"))
            if (
                existing_meta.get("file_name") == file_name
                and existing_meta.get("size") == len(file_bytes)
                and existing_meta.get("sha256") == resolved_hash
            ):
                return
        data_path.write_bytes(file_bytes)
        meta = {
            "last_selected_format": format_name,
            "file_name": file_name,
            "size": len(file_bytes),
            "sha256": resolved_hash,
            "saved_at": datetime.now().isoformat(timespec="seconds"),
        }
        meta_path.write_text(json.dumps(meta), encoding="utf-8")
    except Exception:

        pass


def load_persistent_upload(format_name: str):
    try:
        data_path, meta_path = persistent_upload_paths(format_name)
        if not data_path.exists() or not meta_path.exists():
            return None
        file_bytes = data_path.read_bytes()
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        if meta.get("sha256") != stable_file_hash(file_bytes):
            return None
        return {
            "file_bytes": file_bytes,
            "file_name": meta.get("file_name", "Saved report.xlsx"),
            "size": meta.get("size", len(file_bytes)),
            "sha256": meta.get("sha256", ""),
            "saved_at": meta.get("saved_at", ""),
            "last_selected_format": meta.get("last_selected_format", ""),
        }
    except Exception:
        return None


def find_header_row(raw: pd.DataFrame) -> int:
    for idx in range(len(raw)):
        row_values = [clean_text(x).lower() for x in raw.iloc[idx].tolist()]
        has_sku = "sku" in row_values
        has_activity = "activity date" in row_values
        has_ref = "ref #" in row_values
        if has_sku and has_activity and has_ref:
            return idx
    raise ValueError("Cannot find a header row with SKU / Activity Date / Ref #.")


def validate_selected_format(raw: pd.DataFrame, config: dict):
    header_idx = find_header_row(raw)
    header_row = raw.iloc[header_idx]
    cols = config["cols"]
    required_headers = {
        "sku": "sku",
        "activity_date": "activity date",
        "ref_no": "ref #",
        "qty_in": "qty",
        "qty_out": "qty",
        "balance": "balance",
    }

    for key, expected_text in required_headers.items():
        actual_header = clean_text(get_cell(header_row, cols[key])).lower()
        if expected_text not in actual_header:
            raise WrongFileFormatError(config["wrong_format_warning"])

    return header_idx


def extract_report_range(raw: pd.DataFrame):
    start_dt, end_dt = pd.NaT, pd.NaT

    for r in range(min(15, len(raw))):
        row = raw.iloc[r].tolist()
        row_text_original = " | ".join(clean_text(x) for x in row if clean_text(x))
        row_text = row_text_original.lower()
        if "item activity from" not in row_text:
            continue

        date_matches = re.findall(
            r"\d{1,2}/\d{1,2}/\d{2,4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?\s*(?:AM|PM)?)?",
            row_text_original,
            flags=re.IGNORECASE,
        )
        if len(date_matches) >= 2:
            start_dt = parse_excel_or_text_date(date_matches[0])
            end_dt = parse_excel_or_text_date(date_matches[1])
            break

        match = re.search(
            r"item\s+activity\s+from:\s*(.*?)\s+to\s+([^|]+)",
            row_text_original,
            flags=re.IGNORECASE,
        )
        if match:
            start_dt = parse_excel_or_text_date(match.group(1))
            end_dt = parse_excel_or_text_date(match.group(2))
            break

        to_index = None
        for i, value in enumerate(row):
            if clean_text(value).lower() == "to":
                to_index = i
                break

        dates_before_to, dates_after_to = [], []
        for i, value in enumerate(row):
            parsed = parse_excel_or_text_date(value)
            if not pd.isna(parsed):
                if to_index is not None and i > to_index:
                    dates_after_to.append(parsed)
                else:
                    dates_before_to.append(parsed)

        if dates_before_to:
            start_dt = dates_before_to[0]
        if dates_after_to:
            end_dt = dates_after_to[0]
        elif len(dates_before_to) >= 2:
            end_dt = dates_before_to[1]
        break

    return start_dt, end_dt


@st.cache_data(show_spinner=False)
def load_excel_to_raw(file_bytes: bytes, cache_version: str = APP_CACHE_VERSION) -> pd.DataFrame:
    return pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=None, dtype=object)


@st.cache_data(show_spinner=False)
def process_excel_file(file_bytes: bytes, format_name: str, cache_version: str = APP_CACHE_VERSION) -> dict:
    raw_df = load_excel_to_raw(file_bytes, cache_version)
    config = FORMAT_CONFIGS[format_name]
    validate_selected_format(raw_df, config)
    return build_inventory_model(raw_df, config, format_name)


def last_data_activity_dates(tx_df: pd.DataFrame, end_date, count: int, start_date=None) -> list:
    end_date = pd.to_datetime(end_date).normalize()
    if tx_df is None or tx_df.empty or "Activity Date" not in tx_df.columns:
        valid_dates = pd.DatetimeIndex([])
    else:
        valid_dates = pd.DatetimeIndex(
            pd.to_datetime(tx_df.loc[tx_df["Activity Date"].notna(), "Activity Date"], errors="coerce")
            .dropna()
            .dt.normalize()
            .drop_duplicates()
            .sort_values()
        )
    valid_dates = valid_dates[valid_dates <= end_date]
    if start_date is not None and not pd.isna(start_date):
        valid_dates = valid_dates[valid_dates >= pd.to_datetime(start_date).normalize()]
    if end_date not in valid_dates:
        valid_dates = valid_dates.append(pd.DatetimeIndex([end_date]))
    return valid_dates.sort_values().unique()[-count:].tolist()


def add_business_days(start_date, days):
    if pd.isna(days) or not np.isfinite(days):
        return pd.NaT
    whole_days = max(int(np.ceil(float(days))), 0)
    start = pd.to_datetime(start_date).normalize()
    if whole_days == 0:
        return start
    return start + (whole_days * WAREHOUSE_BUSINESS_DAY)


def build_inventory_model(raw: pd.DataFrame, config: dict, format_name: str) -> dict:
    header_idx = find_header_row(raw)
    report_start, report_end = extract_report_range(raw)
    rows = raw.iloc[header_idx + 1 :]
    cols = config["cols"]

    current_sku = ""
    current_desc = ""
    sku_records = {}
    report_sku_order = []
    transactions = []
    official_total_rows = []
    official_ending_rows = []
    beginning_balance_rows = []
    not_shipped_rows = []
    cancelled_rows = []

    def ensure_sku_record(sku, desc):
        sku_records.setdefault(
            sku,
            {
                "SKU": sku,
                "Description": desc,
                "Official Total Inbound": np.nan,
                "Official Total Outbound": np.nan,
                "Ending Balance": np.nan,
                "Ctn Balance": np.nan,
                "Beginning Balance": np.nan,
                "Official Ending Row": np.nan,
                "Official Total Row": np.nan,
                "Last Activity Date": pd.NaT,
                "Last Inbound Date": pd.NaT,
                "Last Outbound Date": pd.NaT,
            },
        )
        if desc:
            sku_records[sku]["Description"] = desc

    for excel_row_num, row in zip(rows.index, rows.itertuples(index=False, name=None)):
        sku_cell = clean_text(get_cell(row, cols["sku"]))
        desc_cell = clean_text(get_cell(row, cols["description"]))
        activity_raw = get_cell(row, cols["activity_date"])
        activity_text = clean_text(activity_raw)
        ref_text = clean_text(get_cell(row, cols["ref_no"]))
        trans_no = clean_text(get_cell(row, cols["trans_no"]))
        qty_in_raw = clean_text(get_cell(row, cols["qty_in"]))
        qty_out_raw = clean_text(get_cell(row, cols["qty_out"]))
        qty_in_value = first_qty_number(get_cell(row, cols["qty_in"]), np.nan)
        qty_out_value = first_qty_number(get_cell(row, cols["qty_out"]), np.nan)
        balance_value = first_qty_number(get_cell(row, cols["balance"]), np.nan)
        ctn_balance_value = first_qty_number(get_cell(row, cols["ctn_balance"]), np.nan) if cols.get("ctn_balance") is not None else np.nan
        qty_in = 0.0 if pd.isna(qty_in_value) else float(qty_in_value)
        qty_out = 0.0 if pd.isna(qty_out_value) else float(qty_out_value)

        sku_lower = sku_cell.lower()
        activity_lower = activity_text.lower()
        ref_lower = ref_text.lower()
        activity_key = activity_lower.rstrip(":")
        is_total_row = (config["total_rule"] == "sku_totals" and sku_lower.rstrip(":") == "totals") or (config["total_rule"] == "ref_total" and ref_lower.rstrip(":") == "total")

        if sku_cell and sku_lower != "sku" and not is_total_row:
            current_sku = sku_cell
            current_desc = desc_cell
            if current_sku not in sku_records:
                report_sku_order.append(current_sku)
            ensure_sku_record(current_sku, current_desc)
            continue

        if not current_sku:
            continue

        ensure_sku_record(current_sku, current_desc)

        if activity_key == "beginning balance":
            sku_records[current_sku]["Beginning Balance"] = balance_value
            beginning_balance_rows.append(
                {
                    "Excel Row": excel_row_num + 1,
                    "SKU": current_sku,
                    "Description": sku_records[current_sku]["Description"],
                    "Activity Date": activity_text,
                    "Balance": balance_value,
                    "Ctn Balance": ctn_balance_value,
                }
            )
            continue

        if activity_key == "ending balance":
            sku_records[current_sku]["Ending Balance"] = balance_value
            sku_records[current_sku]["Ctn Balance"] = ctn_balance_value
            sku_records[current_sku]["Official Ending Row"] = excel_row_num + 1
            official_ending_rows.append(
                {
                    "Excel Row": excel_row_num + 1,
                    "SKU": current_sku,
                    "Description": sku_records[current_sku]["Description"],
                    "Activity Date": activity_text,
                    "Balance": balance_value,
                    "Ctn Balance": ctn_balance_value,
                }
            )
            continue

        if is_total_row:
            official_inbound = 0.0 if pd.isna(qty_in_value) else float(qty_in_value)
            official_outbound = 0.0 if pd.isna(qty_out_value) else float(qty_out_value)
            sku_records[current_sku]["Official Total Inbound"] = official_inbound
            sku_records[current_sku]["Official Total Outbound"] = official_outbound
            sku_records[current_sku]["Official Total Row"] = excel_row_num + 1
            official_total_rows.append(
                {
                    "Excel Row": excel_row_num + 1,
                    "SKU": current_sku,
                    "Description": sku_records[current_sku]["Description"],
                    "Source": config["total_source"],
                    "Official Total Inbound": official_inbound,
                    "Official Total Outbound": official_outbound,
                    "Balance": balance_value,
                    "Ctn Balance": ctn_balance_value,
                }
            )
            continue

        activity_dt = parse_excel_or_text_date(activity_raw)
        is_cancelled = "cancel" in ref_lower or "cancel" in activity_lower
        is_not_shipped = "not shipped" in activity_lower or "not shipped" in ref_lower

        if is_cancelled:
            cancelled_rows.append(
                {
                    "Excel Row": excel_row_num + 1,
                    "SKU": current_sku,
                    "Description": sku_records[current_sku]["Description"],
                    "Activity Date": activity_dt,
                    "Ref #": ref_text,
                    "Qty Out": qty_out,
                }
            )

        if is_not_shipped:
            not_shipped_rows.append(
                {
                    "Excel Row": excel_row_num + 1,
                    "SKU": current_sku,
                    "Description": sku_records[current_sku]["Description"],
                    "Activity Date": activity_dt,
                    "Ref #": ref_text,
                    "Qty Out": qty_out,
                }
            )

        if not pd.isna(activity_dt):
            has_inbound = qty_in != 0
            has_outbound = qty_out != 0
            if qty_in > 0 and qty_out > 0:
                transaction_type = "Inbound / Outbound"
            elif qty_in > 0:
                transaction_type = "Inbound"
            elif qty_out > 0:
                transaction_type = "Outbound"
            elif qty_in < 0 or qty_out < 0:
                transaction_type = "Adjustment"
            elif is_cancelled:
                transaction_type = "Cancelled / No Qty"
            else:
                transaction_type = "No Qty"

            transactions.append(
                {
                    "Excel Row": excel_row_num + 1,
                    "SKU": current_sku,
                    "Description": sku_records[current_sku]["Description"],
                    "Activity Date": activity_dt,
                    "Transaction Type": transaction_type,
                    "Trans. #": trans_no,
                    "Ref #": ref_text,
                    "Qty In / Ctn Raw": qty_in_raw,
                    "Qty Out / Ctn Raw": qty_out_raw,
                    "Qty In": qty_in,
                    "Qty Out": qty_out,
                    "Balance After Transaction": balance_value,
                    "Ctn Balance After Transaction": ctn_balance_value,
                    "Is Not Shipped": is_not_shipped,
                    "Is Cancelled": is_cancelled,
                }
            )

            if has_inbound or has_outbound:
                existing_last = sku_records[current_sku]["Last Activity Date"]
                if pd.isna(existing_last) or activity_dt > existing_last:
                    sku_records[current_sku]["Last Activity Date"] = activity_dt
            if qty_in > 0:
                existing_inbound = sku_records[current_sku]["Last Inbound Date"]
                if pd.isna(existing_inbound) or activity_dt > existing_inbound:
                    sku_records[current_sku]["Last Inbound Date"] = activity_dt
            if qty_out > 0:
                existing_outbound = sku_records[current_sku]["Last Outbound Date"]
                if pd.isna(existing_outbound) or activity_dt > existing_outbound:
                    sku_records[current_sku]["Last Outbound Date"] = activity_dt

    tx_columns = [
        "Excel Row",
        "SKU",
        "Description",
        "Activity Date",
        "Transaction Type",
        "Trans. #",
        "Ref #",
        "Qty In / Ctn Raw",
        "Qty Out / Ctn Raw",
        "Qty In",
        "Qty Out",
        "Balance After Transaction",
        "Ctn Balance After Transaction",
        "Is Not Shipped",
        "Is Cancelled",
    ]
    sku_df = pd.DataFrame(sku_records.values())
    tx_df = pd.DataFrame(transactions, columns=tx_columns)
    official_total_df = pd.DataFrame(official_total_rows)
    official_ending_df = pd.DataFrame(official_ending_rows)
    beginning_balance_df = pd.DataFrame(beginning_balance_rows)
    not_shipped_df = pd.DataFrame(not_shipped_rows)
    cancelled_df = pd.DataFrame(cancelled_rows)

    if sku_df.empty:
        raise ValueError(f"No SKU sections were found in the {format_name} file.")
    parsed_sku_names = set(sku_df["SKU"].astype(str))
    missing_report_skus = [sku for sku in report_sku_order if sku not in parsed_sku_names]
    if missing_report_skus or len(sku_df) != len(report_sku_order):
        missing_text = ", ".join(missing_report_skus[:10]) or "SKU count mismatch"
        raise ValueError(f"Not all report SKUs were loaded: {missing_text}")

    tx_dates = pd.to_datetime(tx_df["Activity Date"], errors="coerce").dropna() if not tx_df.empty else pd.Series(dtype="datetime64[ns]")
    report_date_source = "File Header"
    if pd.isna(report_end):
        if tx_dates.empty:
            raise ValueError("Cannot determine the report end date from the file header or transaction rows.")
        report_end = tx_dates.max()
        report_date_source = "Latest Transaction Date"
    if pd.isna(report_start):
        report_start = tx_dates.min() if not tx_dates.empty else report_end

    report_end = pd.to_datetime(report_end).normalize()
    report_start = pd.to_datetime(report_start).normalize()
    if report_start > report_end:
        raise ValueError("The report start date is later than the report end date.")

    total_counts = official_total_df.groupby("SKU").size().to_dict() if not official_total_df.empty else {}
    ending_counts = official_ending_df.groupby("SKU").size().to_dict() if not official_ending_df.empty else {}

    data_quality_issues = []
    for _, sku_row in sku_df.iterrows():
        sku = sku_row["SKU"]
        issues = []
        total_count = int(total_counts.get(sku, 0))
        ending_count = int(ending_counts.get(sku, 0))
        if total_count == 0:
            issues.append("Missing Total Row")
        elif total_count > 1:
            issues.append("Duplicate Total Rows")
        if ending_count == 0:
            issues.append("Missing Ending Balance Row")
        elif ending_count > 1:
            issues.append("Duplicate Ending Balance Rows")
        if ending_count > 0 and pd.isna(sku_row["Ending Balance"]):
            issues.append("Missing Ending Balance Value")
        data_quality_issues.append("; ".join(issues))

    sku_df["Data Quality Issue"] = data_quality_issues
    sku_df["Data Quality Status"] = np.where(sku_df["Data Quality Issue"] == "", "OK", "Review")
    sku_df["Official Total Row Count"] = sku_df["SKU"].map(total_counts).fillna(0).astype(int)
    sku_df["Official Ending Row Count"] = sku_df["SKU"].map(ending_counts).fillna(0).astype(int)

    positive_outbound_df = tx_df[tx_df["Qty Out"] > 0] if not tx_df.empty else tx_df
    all_window_dates = last_data_activity_dates(tx_df, report_end, 90, report_start)
    window_dates = {
        "Outbound Last 90 Days": all_window_dates,
        "Outbound Last 30 Days": all_window_dates[-30:],
        "Outbound Last 14 Days": all_window_dates[-14:],
        "Outbound Last 7 Days": all_window_dates[-7:],
    }
    windows = {
        label: (dates[0], dates[-1]) if dates else (pd.NaT, pd.NaT)
        for label, dates in window_dates.items()
    }

    for label, dates in window_dates.items():
        if positive_outbound_df.empty or not dates:
            sku_df[label] = 0.0
        else:
            mask = positive_outbound_df["Activity Date"].isin(dates)
            agg = positive_outbound_df.loc[mask].groupby("SKU")["Qty Out"].sum()
            sku_df[label] = sku_df["SKU"].map(agg).fillna(0.0)

    valid_30d_count = len(window_dates["Outbound Last 30 Days"])
    sku_df["Avg Daily Usage 30D"] = np.where(
        valid_30d_count > 0,
        sku_df["Outbound Last 30 Days"] / valid_30d_count,
        0.0,
    )
    sku_df["Demand Status"] = np.select(
        [
            sku_df["Outbound Last 30 Days"] > 0,
            sku_df["Outbound Last 90 Days"] > 0,
        ],
        ["Active", "Dormant"],
        default="Inactive",
    )
    sku_df["Days Remaining"] = np.nan
    usable_mask = (
        (sku_df["Data Quality Status"] == "OK")
        & (sku_df["Demand Status"] == "Active")
        & (sku_df["Avg Daily Usage 30D"] > 0)
    )
    positive_balance_mask = usable_mask & (sku_df["Ending Balance"] > 0)
    zero_balance_mask = (
        (sku_df["Data Quality Status"] == "OK")
        & (sku_df["Demand Status"] == "Active")
        & (sku_df["Ending Balance"] <= 0)
    )
    sku_df.loc[positive_balance_mask, "Days Remaining"] = (
        sku_df.loc[positive_balance_mask, "Ending Balance"] / sku_df.loc[positive_balance_mask, "Avg Daily Usage 30D"]
    )
    sku_df.loc[zero_balance_mask, "Days Remaining"] = 0.0

    def assign_risk(row):
        if row["Data Quality Status"] != "OK":
            return "Data Issue"
        if row["Demand Status"] == "Inactive":
            return "Inactive / No Demand"
        if row["Demand Status"] == "Dormant":
            return "No Recent Demand"
        if row["Ending Balance"] <= 0:
            return "Critical"
        if row["Days Remaining"] <= 7:
            return "Critical"
        if row["Days Remaining"] <= 14:
            return "Warning"
        if row["Days Remaining"] <= 30:
            return "Watch"
        return "Healthy"

    sku_df["Risk Level"] = sku_df.apply(assign_risk, axis=1)
    sku_df["Recommended Action"] = sku_df["Risk Level"].map(
        {
            "Data Issue": "Review source rows before using this SKU",
            "Critical": "Prepare inbound or reserve stock immediately",
            "Warning": "Confirm inbound ETA and reserve inventory",
            "Watch": "Monitor usage and upcoming orders",
            "Healthy": "No immediate action",
            "No Recent Demand": "Review before replenishment",
            "Inactive / No Demand": "No replenishment required",
        }
    )
    sku_df["Forecast Stockout Date"] = sku_df["Days Remaining"].map(lambda value: add_business_days(report_end, value))

    transaction_totals = (
        tx_df.groupby("SKU", as_index=True).agg(
            **{
                "Transaction Total Inbound": ("Qty In", "sum"),
                "Transaction Total Outbound": ("Qty Out", "sum"),
            }
        )
        if not tx_df.empty
        else pd.DataFrame(columns=["Transaction Total Inbound", "Transaction Total Outbound"])
    )
    sku_df["Transaction Total Inbound"] = sku_df["SKU"].map(transaction_totals.get("Transaction Total Inbound", pd.Series(dtype=float))).fillna(0.0)
    sku_df["Transaction Total Outbound"] = sku_df["SKU"].map(transaction_totals.get("Transaction Total Outbound", pd.Series(dtype=float))).fillna(0.0)
    sku_df["Calculated Ending Balance"] = sku_df["Beginning Balance"] + sku_df["Official Total Inbound"] - sku_df["Official Total Outbound"]
    sku_df["Balance Difference"] = sku_df["Ending Balance"] - sku_df["Calculated Ending Balance"]
    sku_df["Inbound Difference"] = sku_df["Official Total Inbound"] - sku_df["Transaction Total Inbound"]
    sku_df["Outbound Difference"] = sku_df["Official Total Outbound"] - sku_df["Transaction Total Outbound"]

    def assign_audit_status(row):
        if row["Data Quality Status"] != "OK":
            return "Review"
        checks = []
        for field in ["Balance Difference", "Inbound Difference", "Outbound Difference"]:
            value = row[field]
            if not pd.isna(value):
                checks.append(abs(float(value)) < 0.01)
        if not checks:
            return "Not Available"
        if not all(checks):
            return "Review"
        return "Pass" if len(checks) == 3 else "Partial"

    sku_df["Audit Status"] = sku_df.apply(assign_audit_status, axis=1)
    audit_cols = [
        "SKU",
        "Description",
        "Data Quality Status",
        "Data Quality Issue",
        "Beginning Balance",
        "Official Total Inbound",
        "Official Total Outbound",
        "Calculated Ending Balance",
        "Ending Balance",
        "Balance Difference",
        "Transaction Total Inbound",
        "Inbound Difference",
        "Transaction Total Outbound",
        "Outbound Difference",
        "Official Total Row Count",
        "Official Ending Row Count",
        "Audit Status",
    ]
    audit_df = sku_df[audit_cols].copy()

    risk_order = {
        "Data Issue": 0,
        "Critical": 1,
        "Warning": 2,
        "Watch": 3,
        "Healthy": 4,
        "No Recent Demand": 5,
        "Inactive / No Demand": 6,
    }
    sku_df["Risk Sort"] = sku_df["Risk Level"].map(risk_order).fillna(9)
    sku_df = sku_df.sort_values(
        by=["Risk Sort", "Days Remaining", "Outbound Last 14 Days", "Outbound Last 30 Days", "SKU"],
        ascending=[True, True, False, False, True],
        na_position="last",
    ).reset_index(drop=True)

    if positive_outbound_df.empty:
        trend_df = pd.DataFrame(columns=["Activity Date", "Qty Out"])
    else:
        trend_df = positive_outbound_df.groupby("Activity Date", as_index=False)["Qty Out"].sum().sort_values("Activity Date")

    return {
        "format_name": format_name,
        "source_sku_count": len(report_sku_order),
        "sku_df": sku_df,
        "tx_df": tx_df,
        "trend_df": trend_df,
        "audit_df": audit_df,
        "official_total_df": official_total_df,
        "official_ending_df": official_ending_df,
        "beginning_balance_df": beginning_balance_df,
        "not_shipped_df": not_shipped_df,
        "cancelled_df": cancelled_df,
        "report_start": report_start,
        "report_end": report_end,
        "report_date_source": report_date_source,
        "windows": windows,
        "window_dates": window_dates,
        "header_idx": header_idx,
        "config": config,
    }


def fmt_num(value, decimals=0):
    if value is None or pd.isna(value):
        return "-"
    if value == np.inf:
        return "∞"
    return f"{value:,.{decimals}f}"


def fmt_date(value):
    if value is None or pd.isna(value):
        return "-"
    return pd.to_datetime(value).strftime("%m/%d/%Y")


def risk_badge_text(level: str) -> str:
    return clean_text(level)


def html_pills(values, class_name, limit=36):
    cleaned = [str(v).strip() for v in values if str(v).strip()]
    if not cleaned:
        return "<span class='tx-pill-muted'>None</span>"
    shown = cleaned[:limit]
    pills = "".join(f"<span class='{class_name}'>{html.escape(value)}</span>" for value in shown)
    if len(cleaned) > limit:
        pills += f"<span class='tx-pill-muted'>+{len(cleaned) - limit:,} more</span>"
    return pills


def metric_card(label, value, help_text=""):
    label_text = clean_text(label)
    tone = "accent"
    icon_code = "E9D2"
    label_lower = label_text.lower()
    if "total sku" in label_lower:
        icon_code = "E71D"
    elif any(keyword in label_lower for keyword in ["critical", "shortage", "missing"]):
        tone = "danger"
        icon_code = "E7BA"
    elif "warning" in label_lower:
        tone = "warning"
        icon_code = "E814"
    elif "watch" in label_lower:
        tone = "watch"
        icon_code = "E823"
    elif any(keyword in label_lower for keyword in ["qty in", "inbound", "healthy"]):
        tone = "success"
        icon_code = "E73E"
    elif any(keyword in label_lower for keyword in ["balance", "remaining"]):
        tone = "neutral"
        icon_code = "E9D2"
    elif "forecast" in label_lower:
        tone = "neutral"
        icon_code = "E787"
    elif "risk" in label_lower:
        icon_code = "E9D9"
        value_lower = clean_text(value).lower()
        if "critical" in value_lower or "data issue" in value_lower:
            tone = "danger"
        elif "warning" in value_lower:
            tone = "warning"
        elif "watch" in value_lower:
            tone = "watch"
        elif "healthy" in value_lower:
            tone = "success"
        else:
            tone = "neutral"

    st.markdown(
        f"""
        <div class="kpi-card kpi-tone-{tone}">
            <div class="kpi-accent" aria-hidden="true"></div>
            <div class="kpi-topline">
                <div class="kpi-label">{html.escape(label_text)}</div>
                <div class="kpi-icon" aria-hidden="true">&#x{icon_code};</div>
            </div>
            <div class="kpi-value">{html.escape(clean_text(value))}</div>
            <div class="kpi-help">{html.escape(clean_text(help_text))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def tab_page_header(title: str, subtitle: str) -> None:
    st.markdown(
        f"""
        <div class="tab-page-heading">
            <div class="tab-page-title">{html.escape(title)}</div>
            <div class="tab-page-subtitle">{html.escape(subtitle)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def round_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    numeric_cols = out.select_dtypes(include=["number"]).columns
    for col in numeric_cols:
        if str(col) == "Avg Daily Usage 30D":
            out[col] = out[col].round(2)
        else:
            out[col] = out[col].replace(np.inf, np.nan).round(0)
    return out


def format_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    date_keywords = ("date", "activity date", "forecast stockout", "last activity")
    for col in out.columns:
        col_lower = str(col).lower()
        if any(keyword in col_lower for keyword in date_keywords):
            parsed = pd.to_datetime(out[col], errors="coerce")
            if parsed.notna().any():
                out[col] = parsed.dt.strftime("%m/%d/%Y").replace("NaT", "")
    return out


def display_table(df: pd.DataFrame) -> pd.DataFrame:
    return format_date_columns(round_numeric_columns(df))


def prepare_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "Risk Level" in out.columns:
        out["Risk Level"] = out["Risk Level"].map(risk_badge_text)
    integer_metric_cols = [
        "Beginning Balance",
        "Ending Balance",
        "Ctn Balance",
        "Official Total Inbound",
        "Official Total Outbound",
        "Outbound Last 90 Days",
        "Outbound Last 30 Days",
        "Outbound Last 14 Days",
        "Outbound Last 7 Days",
        "Official Total Row",
        "Official Ending Row",
        "Official Total Row Count",
        "Official Ending Row Count",
    ]
    for col in integer_metric_cols:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").round(0).astype("Int64")
    if "Avg Daily Usage 30D" in out.columns:
        out["Avg Daily Usage 30D"] = pd.to_numeric(out["Avg Daily Usage 30D"], errors="coerce").round(2)
    if "Days Remaining" in out.columns:
        out["Days Remaining"] = pd.to_numeric(out["Days Remaining"], errors="coerce").round(1)
    for col in ["Forecast Stockout Date", "Last Activity Date", "Last Inbound Date", "Last Outbound Date"]:
        if col in out.columns:
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%m/%d/%Y").replace("NaT", "")
    return out


def prepare_customer_export(df: pd.DataFrame) -> pd.DataFrame:
    customer_cols = [
        "SKU",
        "Description",
        "Risk Level",
        "Ending Balance",
        "Last Outbound Date",
        "Avg Daily Usage 30D",
        "Days Remaining",
        "Forecast Stockout Date",
    ]
    out = df[customer_cols].copy()
    out["Ending Balance"] = pd.to_numeric(out["Ending Balance"], errors="coerce").round(0)
    out["Avg Daily Usage 30D"] = pd.to_numeric(out["Avg Daily Usage 30D"], errors="coerce").round(2)
    out["Days Remaining"] = pd.to_numeric(out["Days Remaining"], errors="coerce").round(1)
    for col in ["Forecast Stockout Date", "Last Outbound Date"]:
        out[col] = pd.to_datetime(out[col], errors="coerce")
    return out


def report_download_filename(format_name: str, report_end) -> str:
    try:
        date_part = pd.to_datetime(report_end).strftime("%m%d%Y")
    except Exception:
        date_part = datetime.today().strftime("%m%d%Y")
    clean_format = re.sub(r"[^A-Za-z0-9]+", "_", str(format_name)).strip("_") or "Inventory"
    return f"{clean_format}_Inventory_Shortage_Report_{date_part}.xlsx"

def prepare_transaction_export(tx_df: pd.DataFrame) -> pd.DataFrame:
    export_cols = [
        "SKU",
        "Description",
        "Activity Date",
        "Transaction Type",
        "Trans. #",
        "Ref #",
        "Qty In",
        "Qty Out",
        "Balance After Transaction",
        "Is Not Shipped",
        "Is Cancelled",
    ]
    out = tx_df.copy()
    for col in export_cols:
        if col not in out.columns:
            if col in ["Qty In", "Qty Out", "Balance After Transaction"]:
                out[col] = 0.0
            elif col in ["Is Not Shipped", "Is Cancelled"]:
                out[col] = False
            elif col == "Activity Date":
                out[col] = pd.NaT
            else:
                out[col] = ""

    out = out[export_cols].copy()
    out["Activity Date"] = pd.to_datetime(out["Activity Date"], errors="coerce")
    for col in ["Qty In", "Qty Out", "Balance After Transaction"]:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).round(0)

    out = out.sort_values(["SKU", "Activity Date"], ascending=[True, False]).reset_index(drop=True)
    return out



def split_stock_input_line(line: str) -> list:
    value = clean_text(line)
    if not value:
        return []
    if "\t" in value:
        parts = value.split("\t")
    elif "|" in value:
        parts = value.split("|")
    elif "," in value:
        parts = value.split(",")
    else:
        parts = re.split(r"\s+", value, maxsplit=2)
    return [clean_text(part) for part in parts]


def parse_stock_check_input(raw_text: str) -> pd.DataFrame:
    rows = []
    order = 0
    for raw_line in str(raw_text or "").splitlines():
        line = clean_text(raw_line)
        if not line:
            continue
        parts = split_stock_input_line(line)
        if len(parts) < 3:
            order += 1
            rows.append(
                {
                    "Input Order": order,
                    "DO #": parts[0] if len(parts) > 0 else "",
                    "SKU": parts[1] if len(parts) > 1 else "",
                    "Requested Qty": np.nan,
                    "Issue": "Missing required columns",
                }
            )
            continue
        header_text = " ".join(parts[:3]).lower()
        if ("do" in header_text or "order" in header_text) and ("sku" in header_text or "item" in header_text) and "qty" in header_text:
            continue
        order += 1
        do_no = clean_text(parts[0])
        sku = clean_text(parts[1])
        qty_text = clean_text(parts[2])
        match = re.search(r"-?\d+(?:\.\d+)?", qty_text.replace(",", ""))
        qty = float(match.group()) if match else np.nan
        issue = ""
        if not do_no:
            issue = "Missing DO #"
        elif not sku:
            issue = "Missing SKU"
        elif pd.isna(qty) or qty <= 0:
            issue = "Invalid Qty"
        rows.append(
            {
                "Input Order": order,
                "DO #": do_no,
                "SKU": sku,
                "Requested Qty": qty,
                "Issue": issue,
            }
        )
    return pd.DataFrame(rows, columns=["Input Order", "DO #", "SKU", "Requested Qty", "Issue"])


def split_stock_check_column(raw_text: str) -> list:
    values = [clean_text(line) for line in str(raw_text or "").splitlines()]
    while values and values[-1] == "":
        values.pop()
    return values


def parse_stock_check_columns(do_text: str, sku_text: str, qty_text: str) -> tuple[pd.DataFrame, dict]:
    do_values = split_stock_check_column(do_text)
    sku_values = split_stock_check_column(sku_text)
    qty_values = split_stock_check_column(qty_text)
    counts = {"DO #": len(do_values), "SKU": len(sku_values), "Qty": len(qty_values)}
    max_rows = max(counts.values()) if counts else 0
    rows = []
    for idx in range(max_rows):
        do_no = clean_text(do_values[idx]) if idx < len(do_values) else ""
        sku = clean_text(sku_values[idx]) if idx < len(sku_values) else ""
        qty_text_value = clean_text(qty_values[idx]) if idx < len(qty_values) else ""
        match = re.search(r"-?\d+(?:\.\d+)?", qty_text_value.replace(",", ""))
        qty = float(match.group()) if match else np.nan
        issue = ""
        if not do_no:
            issue = "Missing DO #"
        elif not sku:
            issue = "Missing SKU"
        elif pd.isna(qty) or qty <= 0:
            issue = "Invalid Qty"
        rows.append(
            {
                "Input Order": idx + 1,
                "DO #": do_no,
                "SKU": sku,
                "Requested Qty": qty,
                "Issue": issue,
            }
        )
    return pd.DataFrame(rows, columns=["Input Order", "DO #", "SKU", "Requested Qty", "Issue"]), counts


def parse_stock_check_table(table_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    if table_df is None or table_df.empty:
        return pd.DataFrame(columns=["Input Order", "DO #", "SKU", "Requested Qty", "Issue"])
    for idx, row in table_df.reset_index(drop=True).iterrows():
        do_no = clean_text(row.get("DO #", ""))
        sku = clean_text(row.get("Item Code / SKU", row.get("SKU", "")))
        qty_text_value = clean_text(row.get("Qty", ""))
        if not do_no and not sku and not qty_text_value:
            continue
        match = re.search(r"-?\d+(?:\.\d+)?", qty_text_value.replace(",", ""))
        qty = float(match.group()) if match else np.nan
        issue = ""
        if not do_no:
            issue = "Missing DO #"
        elif not sku:
            issue = "Missing SKU"
        elif pd.isna(qty) or qty <= 0:
            issue = "Invalid Qty"
        rows.append(
            {
                "Input Order": len(rows) + 1,
                "DO #": do_no,
                "SKU": sku,
                "Requested Qty": qty,
                "Issue": issue,
            }
        )
    return pd.DataFrame(rows, columns=["Input Order", "DO #", "SKU", "Requested Qty", "Issue"])


def exclude_rma_stock_check_rows(input_df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    if input_df is None or input_df.empty or "DO #" not in input_df.columns:
        return input_df.copy() if isinstance(input_df, pd.DataFrame) else pd.DataFrame(), []
    rma_mask = input_df["DO #"].astype(str).str.contains("RMA", case=False, regex=False, na=False)
    excluded_dos = []
    seen_dos = set()
    for value in input_df.loc[rma_mask, "DO #"]:
        do_no = clean_text(value)
        do_key = do_no.upper()
        if do_no and do_key not in seen_dos:
            excluded_dos.append(do_no)
            seen_dos.add(do_key)
    return input_df.loc[~rma_mask].copy(), excluded_dos


def normalize_lookup_key(value) -> str:
    return clean_text(value).upper()


def build_existing_do_tables(tx_df: pd.DataFrame) -> tuple[set, dict, dict]:
    if tx_df is None or tx_df.empty:
        return set(), {}, {}

    source_rows = []
    source_data_cols = ["SKU", "Description", "Qty Out", "Qty In", "Activity Date", "Excel Row"]
    for source_col in ["Ref #", "Trans. #"]:
        if source_col not in tx_df.columns:
            continue
        source_values = tx_df[source_col].astype(str)
        do_keys = source_values.map(normalize_lookup_key)
        valid_mask = do_keys != ""
        if not valid_mask.any():
            continue
        available_cols = [col for col in source_data_cols if col in tx_df.columns]
        tmp = tx_df.loc[valid_mask, available_cols].copy()
        for col in source_data_cols:
            if col not in tmp.columns:
                tmp[col] = np.nan
        tmp["DO Key"] = do_keys.loc[valid_mask].to_numpy()
        if tmp.empty:
            continue
        tmp["DO #"] = source_values.loc[valid_mask].map(clean_text).to_numpy()
        source_rows.append(tmp)

    if not source_rows:
        return set(), {}, {}

    all_do_df = pd.concat(source_rows, ignore_index=True, copy=False).drop_duplicates(subset=["DO Key", "SKU", "Excel Row"])
    all_do_df["SKU Key"] = all_do_df["SKU"].astype(str).map(normalize_lookup_key)
    all_do_df["Qty Out"] = pd.to_numeric(all_do_df.get("Qty Out", 0), errors="coerce").fillna(0)
    all_do_df["Qty In"] = pd.to_numeric(all_do_df.get("Qty In", 0), errors="coerce").fillna(0)
    all_do_df["Activity Date"] = pd.to_datetime(all_do_df.get("Activity Date", pd.NaT), errors="coerce")

    item_df = (
        all_do_df.groupby(["DO Key", "SKU Key"], sort=False, as_index=False)
        .agg(
            {
                "DO #": "first",
                "SKU": "first",
                "Description": "first",
                "Qty Out": "sum",
                "Qty In": "sum",
                "Activity Date": "max",
            }
        )
    )
    do_df = (
        all_do_df.groupby("DO Key", sort=False, as_index=False)
        .agg(
            {
                "DO #": "first",
                "Qty Out": "sum",
                "Qty In": "sum",
                "Activity Date": "max",
                "SKU Key": "nunique",
            }
        )
    )

    existing_do_keys = set(do_df["DO Key"].astype(str))
    existing_item_lookup = item_df.set_index(["DO Key", "SKU Key"]).to_dict("index")
    existing_do_lookup = do_df.set_index("DO Key").to_dict("index")
    return existing_do_keys, existing_item_lookup, existing_do_lookup


def build_stock_check_tables(input_df: pd.DataFrame, sku_df: pd.DataFrame, tx_df: pd.DataFrame, existing_do_tables=None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if input_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    issues_df = input_df[input_df["Issue"].astype(str) != ""].copy()
    valid_df = input_df[input_df["Issue"].astype(str) == ""].copy()
    if valid_df.empty:
        return pd.DataFrame(), pd.DataFrame(), issues_df

    valid_df["SKU Key"] = valid_df["SKU"].astype(str).map(normalize_lookup_key)
    valid_df["DO Key"] = valid_df["DO #"].astype(str).map(normalize_lookup_key)
    request_df = (
        valid_df.groupby(["DO Key", "DO #", "SKU Key", "SKU"], sort=False, as_index=False)
        .agg({"Input Order": "min", "Requested Qty": "sum"})
        .sort_values("Input Order")
        .reset_index(drop=True)
    )

    stock_df = sku_df[["SKU", "Description", "Ending Balance"]].copy()
    stock_df["SKU Key"] = stock_df["SKU"].astype(str).map(normalize_lookup_key)
    stock_df = stock_df.drop_duplicates("SKU Key", keep="first")
    stock_lookup = stock_df.set_index("SKU Key").to_dict("index")
    if existing_do_tables is None:
        existing_do_keys, existing_item_lookup, existing_do_lookup = build_existing_do_tables(tx_df)
    else:
        existing_do_keys, existing_item_lookup, existing_do_lookup = existing_do_tables
    remaining = {}
    detail_rows = []

    for _, row in request_df.iterrows():
        do_key = row["DO Key"]
        sku_key = row["SKU Key"]
        requested_qty = float(row["Requested Qty"])
        stock_info = stock_lookup.get(sku_key)
        existing_item = existing_item_lookup.get((do_key, sku_key), {})
        existing_report_qty = float(pd.to_numeric(pd.Series([existing_item.get("Qty Out", 0)]), errors="coerce").fillna(0).iloc[0])
        qty_to_check = max(0.0, requested_qty - existing_report_qty)
        report_do_status = "Existing DO" if do_key in existing_do_keys else "New DO"
        report_item_status = "Found" if existing_item else ("Not Found" if report_do_status == "Existing DO" else "New")

        if stock_info is None:
            description = existing_item.get("Description", "")
            status = "Already Covered" if report_do_status == "Existing DO" and qty_to_check <= 0 else "Not Found"
            detail_rows.append(
                {
                    "Input Order": int(row["Input Order"]),
                    "DO #": row["DO #"],
                    "SKU": existing_item.get("SKU", row["SKU"]),
                    "Description": description,
                    "Report DO Status": report_do_status,
                    "Report Item Status": report_item_status,
                    "Current Stock": np.nan,
                    "Available Before": np.nan,
                    "Requested Qty": requested_qty,
                    "Existing Report Qty Out": existing_report_qty,
                    "Qty To Check": qty_to_check,
                    "Remaining After This Check": np.nan,
                    "Shortage Qty": qty_to_check if status == "Not Found" else 0.0,
                    "Status": status,
                }
            )
            continue

        current_stock_value = pd.to_numeric(pd.Series([stock_info.get("Ending Balance", np.nan)]), errors="coerce").iloc[0]
        if pd.isna(current_stock_value):
            detail_rows.append(
                {
                    "Input Order": int(row["Input Order"]),
                    "DO #": row["DO #"],
                    "SKU": stock_info.get("SKU", row["SKU"]),
                    "Description": stock_info.get("Description", existing_item.get("Description", "")),
                    "Report DO Status": report_do_status,
                    "Report Item Status": report_item_status,
                    "Current Stock": np.nan,
                    "Available Before": np.nan,
                    "Requested Qty": requested_qty,
                    "Existing Report Qty Out": existing_report_qty,
                    "Qty To Check": qty_to_check,
                    "Remaining After This Check": np.nan,
                    "Shortage Qty": np.nan,
                    "Status": "Data Issue",
                }
            )
            continue
        current_stock = float(current_stock_value)
        if sku_key not in remaining:
            remaining[sku_key] = current_stock
        available_before = remaining[sku_key]

        if qty_to_check <= 0:
            remaining_after = available_before
            shortage_qty = 0.0
            status = "Already Covered" if report_do_status == "Existing DO" else "Enough"
        else:
            remaining_after = available_before - qty_to_check
            shortage_qty = max(0.0, qty_to_check - max(available_before, 0.0))
            status = "Enough" if remaining_after >= 0 else "Shortage"
            remaining[sku_key] = remaining_after

        detail_rows.append(
            {
                "Input Order": int(row["Input Order"]),
                "DO #": row["DO #"],
                "SKU": stock_info.get("SKU", row["SKU"]),
                "Description": stock_info.get("Description", existing_item.get("Description", "")),
                "Report DO Status": report_do_status,
                "Report Item Status": report_item_status,
                "Current Stock": current_stock,
                "Available Before": available_before,
                "Requested Qty": requested_qty,
                "Existing Report Qty Out": existing_report_qty,
                "Qty To Check": qty_to_check,
                "Remaining After This Check": remaining_after,
                "Shortage Qty": shortage_qty,
                "Status": status,
            }
        )

    detail_df = pd.DataFrame(detail_rows)
    overview_rows = []
    do_order_df = valid_df[["DO Key", "DO #", "Input Order"]].drop_duplicates("DO Key").sort_values("Input Order")
    for _, do_row in do_order_df.iterrows():
        do_key = do_row["DO Key"]
        do_detail = detail_df[detail_df["DO #"].astype(str).map(normalize_lookup_key) == do_key].copy()
        if do_detail.empty:
            continue
        shortage_items = int((do_detail["Status"] == "Shortage").sum())
        not_found_items = int((do_detail["Status"] == "Not Found").sum())
        data_issue_items = int((do_detail["Status"] == "Data Issue").sum())
        enough_items = int((do_detail["Status"] == "Enough").sum())
        covered_items = int((do_detail["Status"] == "Already Covered").sum())
        if data_issue_items > 0:
            status = "Data Issue"
        elif shortage_items > 0:
            status = "Shortage"
        elif not_found_items > 0:
            status = "Not Found"
        elif enough_items > 0:
            status = "Enough"
        else:
            status = "Already Covered"
        existing_do_info = existing_do_lookup.get(do_key, {})
        overview_rows.append(
            {
                "DO #": do_row["DO #"],
                "Report DO Status": "Existing DO" if do_key in existing_do_keys else "New DO",
                "Status": status,
                "Item Count": len(do_detail),
                "Total Requested Qty": pd.to_numeric(do_detail["Requested Qty"], errors="coerce").fillna(0).sum(),
                "Existing Report Qty Out": pd.to_numeric(do_detail["Existing Report Qty Out"], errors="coerce").fillna(0).sum(),
                "Qty To Check": pd.to_numeric(do_detail["Qty To Check"], errors="coerce").fillna(0).sum(),
                "Enough Items": enough_items,
                "Already Covered Items": covered_items,
                "Shortage Items": shortage_items,
                "Not Found Items": not_found_items,
                "Data Issue Items": data_issue_items,
                "Report Activity Date": existing_do_info.get("Activity Date", pd.NaT),
            }
        )

    overview_df = pd.DataFrame(overview_rows)
    return detail_df, overview_df, issues_df


def stock_status_badge(value: str) -> str:
    return {
        "Enough": "Enough",
        "Shortage": "⚠️ Shortage",
        "Not Found": "Missing SKU",
        "Invalid Qty": "Invalid Qty",
        "Already Covered": "Already in Report",
        "No Change": "No Change",
        "Data Issue": "Data Issue",
    }.get(str(value), str(value))


def prepare_stock_check_display(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "Status" in out.columns:
        out["Status"] = out["Status"].map(stock_status_badge)
    if "Temporary Status" in out.columns:
        out["Temporary Status"] = out["Temporary Status"].map(stock_status_badge)
    numeric_cols = [
        "Current Stock",
        "Current Ending Balance",
        "Available Before",
        "Requested Qty",
        "Existing Report Qty Out",
        "Qty To Check",
        "Total Qty To Check",
        "Remaining After This Check",
        "Remaining After This DO",
        "Temporary Balance",
        "Shortage Qty",
        "Total Requested Qty",
    ]
    for col in numeric_cols:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").round(0).astype("Int64")
    count_cols = ["Item Count", "Enough Items", "Already Covered Items", "Shortage Items", "Not Found Items", "Data Issue Items"]
    for col in count_cols:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0).round(0).astype("Int64")
    for date_col in ["Report Activity Date", "Last Activity Date"]:
        if date_col in out.columns:
            out[date_col] = pd.to_datetime(out[date_col], errors="coerce").dt.strftime("%m/%d/%Y").replace("NaT", "")
    return out


def build_temporary_balance_table(sku_df: pd.DataFrame, detail_df: pd.DataFrame) -> pd.DataFrame:
    if sku_df.empty:
        return pd.DataFrame()

    base_cols = ["SKU", "Description", "Ending Balance", "Risk Level", "Last Activity Date"]
    base = sku_df.copy()
    for col in base_cols:
        if col not in base.columns:
            base[col] = "" if col not in ["Ending Balance", "Last Activity Date"] else np.nan

    balance_df = base[base_cols].copy()
    balance_df["SKU Key"] = balance_df["SKU"].astype(str).map(normalize_lookup_key)
    balance_df["Current Ending Balance"] = pd.to_numeric(balance_df["Ending Balance"], errors="coerce")

    if detail_df.empty or "Qty To Check" not in detail_df.columns:
        impact_df = pd.DataFrame(columns=["SKU Key", "Total Qty To Check", "Affected DO #"])
    else:
        impact_source = detail_df.copy()
        impact_source["SKU Key"] = impact_source["SKU"].astype(str).map(normalize_lookup_key)
        impact_source["Qty To Check"] = pd.to_numeric(impact_source["Qty To Check"], errors="coerce").fillna(0)
        impact_source = impact_source[(impact_source["SKU Key"] != "") & (impact_source["Qty To Check"] > 0)].copy()
        if impact_source.empty:
            impact_df = pd.DataFrame(columns=["SKU Key", "Total Qty To Check", "Affected DO #"])
        else:
            impact_df = (
                impact_source.groupby("SKU Key", sort=False)
                .agg(
                    **{
                        "Total Qty To Check": ("Qty To Check", "sum"),
                        "Affected DO #": ("DO #", lambda values: ", ".join(dict.fromkeys(clean_text(v) for v in values if clean_text(v)))),
                    }
                )
                .reset_index()
            )

    balance_df = balance_df.merge(impact_df, on="SKU Key", how="left")
    balance_df["Total Qty To Check"] = pd.to_numeric(balance_df["Total Qty To Check"], errors="coerce").fillna(0)
    balance_df["Affected DO #"] = balance_df["Affected DO #"].fillna("")
    balance_df["Temporary Balance"] = balance_df["Current Ending Balance"] - balance_df["Total Qty To Check"]
    balance_df["Shortage Qty"] = np.where(
        balance_df["Current Ending Balance"].isna(),
        np.nan,
        np.where(balance_df["Temporary Balance"] < 0, balance_df["Temporary Balance"].abs(), 0),
    )
    balance_df["Temporary Status"] = np.select(
        [
            balance_df["Current Ending Balance"].isna(),
            balance_df["Total Qty To Check"] <= 0,
            balance_df["Temporary Balance"] >= 0,
        ],
        ["Data Issue", "No Change", "Enough"],
        default="Shortage",
    )
    status_sort = {"Data Issue": 0, "Shortage": 1, "Enough": 2, "No Change": 3}
    balance_df["Impact Sort"] = np.where(balance_df["Total Qty To Check"] > 0, 0, 1)
    balance_df["Status Sort"] = balance_df["Temporary Status"].map(status_sort).fillna(9)
    balance_df = balance_df.sort_values(
        ["Impact Sort", "Status Sort", "Temporary Balance", "SKU"],
        ascending=[True, True, True, True],
    ).reset_index(drop=True)
    if len(balance_df) != len(sku_df):
        raise ValueError("Not all report SKUs were included in the temporary balance.")

    return balance_df[
        [
            "SKU",
            "Description",
            "Risk Level",
            "Current Ending Balance",
            "Total Qty To Check",
            "Temporary Balance",
            "Shortage Qty",
            "Temporary Status",
            "Affected DO #",
            "Last Activity Date",
            "Impact Sort",
            "Status Sort",
        ]
    ]


def transaction_download_filename(format_name: str, report_end) -> str:
    try:
        date_part = pd.to_datetime(report_end).strftime("%m%d%Y")
    except Exception:
        date_part = datetime.today().strftime("%m%d%Y")
    clean_format = re.sub(r"[^A-Za-z0-9]+", "_", str(format_name)).strip("_") or "Inventory"
    return f"{clean_format}_Full_Transaction_History_{date_part}.xlsx"


@st.cache_data(show_spinner=False)
def to_transaction_excel_bytes(model: dict, format_name: str, cache_version: str = APP_CACHE_VERSION) -> bytes:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    report_end = model.get("report_end")
    export_df = prepare_transaction_export(model.get("tx_df", pd.DataFrame()))

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Transaction History", startrow=4, index=False)
        worksheet = writer.sheets["Transaction History"]

        last_col = max(export_df.shape[1], 1)
        last_row = len(export_df) + 5
        header_row = 5

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "Full Transaction History"
        title_cell.font = Font(bold=True, size=18, color="111827")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 26

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        range_cell = worksheet.cell(row=2, column=1)
        range_cell.value = f"Report Date: {fmt_date(report_end)}"
        range_cell.font = Font(size=10, color="4B5563")
        range_cell.alignment = Alignment(horizontal="left", vertical="center")

        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF")
        thin_gray = Side(style="thin", color="E5E7EB")
        border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

        for cell in worksheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        worksheet.row_dimensions[header_row].height = 28

        integer_columns = {"Qty In", "Qty Out", "Balance After Transaction"}
        date_columns = {"Activity Date"}
        text_columns = {"SKU", "Description", "Transaction Type", "Trans. #", "Ref #", "Is Not Shipped", "Is Cancelled"}

        type_styles = {
            "Inbound": {"fill": "DFF3E3", "font": "067647"},
            "Outbound": {"fill": "FDE2E1", "font": "B42318"},
            "Inbound / Outbound": {"fill": "E0F2FE", "font": "026AA2"},
            "Adjustment": {"fill": "F3F4F6", "font": "4B5563"},
            "Cancelled / No Qty": {"fill": "F3F4F6", "font": "4B5563"},
            "No Qty": {"fill": "F3F4F6", "font": "4B5563"},
        }
        tx_type_col_idx = list(export_df.columns).index("Transaction Type") + 1 if "Transaction Type" in export_df.columns else None

        for row in worksheet.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=1, max_col=last_col):
            worksheet.row_dimensions[row[0].row].height = 22
            for cell in row:
                header = worksheet.cell(row=header_row, column=cell.column).value
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="left" if header in text_columns else "right",
                    vertical="center",
                    wrap_text=header in {"Description"},
                )
                if header in integer_columns:
                    cell.number_format = "#,##0"
                elif header in date_columns:
                    cell.number_format = "mm/dd/yyyy"

            if tx_type_col_idx:
                tx_cell = row[tx_type_col_idx - 1]
                style = type_styles.get(str(tx_cell.value), None)
                if style:
                    tx_cell.fill = PatternFill("solid", fgColor=style["fill"])
                    tx_cell.font = Font(bold=True, color=style["font"])
                    tx_cell.alignment = Alignment(horizontal="center", vertical="center")

        preferred_widths = {
            "SKU": 24,
            "Description": 42,
            "Activity Date": 16,
            "Transaction Type": 20,
            "Trans. #": 18,
            "Ref #": 24,
            "Qty In": 14,
            "Qty Out": 14,
            "Balance After Transaction": 24,
            "Is Not Shipped": 16,
            "Is Cancelled": 14,
        }
        for idx, col_name in enumerate(export_df.columns, start=1):
            worksheet.column_dimensions[get_column_letter(idx)].width = preferred_widths.get(col_name, 16)

        worksheet.freeze_panes = "A6"
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5

    return output.getvalue()


@st.cache_data(show_spinner=False)
def to_excel_bytes(model: dict, format_name: str, export_version: str) -> bytes:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    report_start = model.get("report_start")
    report_end = model.get("report_end")
    export_df = prepare_customer_export(model["sku_df"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Shortage Priority", startrow=4, index=False)
        worksheet = writer.sheets["Shortage Priority"]

        last_col = export_df.shape[1]
        last_row = len(export_df) + 5
        header_row = 5


        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = "Inventory Status Report"
        title_cell.font = Font(bold=True, size=18, color="111827")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 26

        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
        range_cell = worksheet.cell(row=2, column=1)
        range_cell.value = f"Report Range: {fmt_date(report_start)} - {fmt_date(report_end)}"
        range_cell.font = Font(size=10, color="4B5563")
        range_cell.alignment = Alignment(horizontal="left", vertical="center")

        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
        note_cell = worksheet.cell(row=3, column=1)
        note_cell.value = "Forecast Stockout Date excludes weekends and U.S. federal holidays."
        note_cell.font = Font(size=10, italic=True, color="6B7280")
        note_cell.alignment = Alignment(horizontal="left", vertical="center")


        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF")
        thin_gray = Side(style="thin", color="E5E7EB")
        border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        body_alignment = Alignment(vertical="center", wrap_text=False)
        text_alignment = Alignment(vertical="center", wrap_text=True)

        risk_styles = {
            "Critical": {"fill": "FDE2E1", "font": "B42318"},
            "Warning": {"fill": "FFE7C2", "font": "B54708"},
            "Watch": {"fill": "FEF7C3", "font": "854D0E"},
            "Healthy": {"fill": "DFF3E3", "font": "067647"},
            "No Recent Demand": {"fill": "F3F4F6", "font": "4B5563"},
            "Inactive / No Demand": {"fill": "F3F4F6", "font": "6B7280"},
            "Data Issue": {"fill": "E5E7EB", "font": "111827"},
        }


        for cell in worksheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        worksheet.row_dimensions[header_row].height = 28


        risk_col_idx = list(export_df.columns).index("Risk Level") + 1
        date_columns = {"Forecast Stockout Date", "Last Activity Date", "Last Outbound Date"}
        decimal_columns = {"Avg Daily Usage 30D", "Days Remaining"}
        integer_columns = {
            "Ending Balance",
            "Official Total Outbound",
            "Outbound Last 90 Days",
            "Outbound Last 30 Days",
            "Outbound Last 14 Days",
            "Outbound Last 7 Days",
        }
        text_columns = {"SKU", "Description", "Risk Level", "Recommended Action", "Demand Status", "Data Quality Issue"}

        for row in worksheet.iter_rows(min_row=header_row + 1, max_row=last_row, min_col=1, max_col=last_col):
            worksheet.row_dimensions[row[0].row].height = 22
            for cell in row:
                header = worksheet.cell(row=header_row, column=cell.column).value
                cell.border = border
                cell.alignment = text_alignment if header in text_columns else body_alignment
                if header in integer_columns:
                    cell.number_format = "#,##0"
                elif header in decimal_columns:
                    cell.number_format = "#,##0.0" if header == "Days Remaining" else "#,##0.00"
                elif header in date_columns:
                    cell.number_format = "mm/dd/yyyy"

            risk_cell = row[risk_col_idx - 1]
            style = risk_styles.get(str(risk_cell.value), None)
            if style:
                risk_cell.fill = PatternFill("solid", fgColor=style["fill"])
                risk_cell.font = Font(bold=True, color=style["font"])
                risk_cell.alignment = Alignment(horizontal="center", vertical="center")


        preferred_widths = {
            "SKU": 24,
            "Description": 42,
            "Risk Level": 16,
            "Recommended Action": 42,
            "Demand Status": 16,
            "Data Quality Issue": 34,
            "Ending Balance": 16,
            "Last Outbound Date": 18,
            "Last Activity Date": 18,
            "Official Total Outbound": 22,
            "Outbound Last 90 Days": 22,
            "Outbound Last 30 Days": 22,
            "Outbound Last 14 Days": 22,
            "Outbound Last 7 Days": 20,
            "Avg Daily Usage 30D": 22,
            "Days Remaining": 18,
            "Forecast Stockout Date": 22,
        }
        for idx, col_name in enumerate(export_df.columns, start=1):
            letter = get_column_letter(idx)
            worksheet.column_dimensions[letter].width = preferred_widths.get(col_name, 16)


        worksheet.freeze_panes = "A6"
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
        worksheet.sheet_view.showGridLines = False

        worksheet.sheet_view.showRowColHeaders = False
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5

    return output.getvalue()

def show_limited_dataframe(df: pd.DataFrame, height: int = 420, limit: int = 500, show_count: bool = True):
    total_rows = len(df)
    if show_count:
        if total_rows > limit:
            st.caption(f"Showing first {limit:,} rows out of {total_rows:,}. Use the export for the complete dataset.")
        else:
            st.caption(f"{total_rows:,} rows")

    display_df = display_table(df.head(limit))
    style_df = pd.DataFrame("", index=display_df.index, columns=display_df.columns)

    if "Risk Level" in display_df.columns:
        risk_styles = {
            "Data Issue": "background-color:#F0F0F0;color:#323130;font-weight:650;text-align:center;",
            "Critical": "background-color:#FDE7E9;color:#A4262C;font-weight:650;text-align:center;",
            "Warning": "background-color:#FFF4CE;color:#8A4B08;font-weight:650;text-align:center;",
            "Watch": "background-color:#FFF8D6;color:#6B5900;font-weight:650;text-align:center;",
            "Healthy": "background-color:#DFF6DD;color:#0B6A0B;font-weight:650;text-align:center;",
            "No Recent Demand": "background-color:#F3F2F1;color:#605E5C;font-weight:600;text-align:center;",
            "Inactive / No Demand": "background-color:#F3F2F1;color:#797775;font-weight:600;text-align:center;",
        }
        for idx, value in display_df["Risk Level"].items():
            style_df.at[idx, "Risk Level"] = risk_styles.get(clean_text(value), "")

    if "Audit Status" in display_df.columns:
        audit_styles = {
            "Pass": "background-color:#DFF6DD;color:#0B6A0B;font-weight:650;text-align:center;",
            "Review": "background-color:#FDE7E9;color:#A4262C;font-weight:650;text-align:center;",
            "Partial": "background-color:#FFF4CE;color:#8A4B08;font-weight:650;text-align:center;",
            "Not Available": "background-color:#F3F2F1;color:#605E5C;font-weight:600;text-align:center;",
        }
        for idx, value in display_df["Audit Status"].items():
            style_df.at[idx, "Audit Status"] = audit_styles.get(clean_text(value), "")

    styled_df = display_df.style.apply(lambda _: style_df, axis=None)
    st.dataframe(styled_df, use_container_width=True, hide_index=True, height=height)


def show_temporary_balance_dataframe(df: pd.DataFrame, height: int = 420, limit: int = 2000, show_count: bool = True):
    total_rows = len(df)
    if show_count:
        if total_rows > limit:
            st.caption(f"Showing first {limit:,} rows out of {total_rows:,} rows for faster loading. Download export for full data.")
        else:
            st.caption(f"Showing {total_rows:,} rows.")

    display_df = prepare_stock_check_display(df.head(limit))

    style_df = pd.DataFrame("", index=display_df.index, columns=display_df.columns)
    if "Temporary Balance" in display_df.columns:
        balance_values = pd.to_numeric(display_df["Temporary Balance"], errors="coerce")
        qty_source = display_df["Total Qty To Check"] if "Total Qty To Check" in display_df.columns else pd.Series(0, index=display_df.index)
        qty_values = pd.to_numeric(qty_source, errors="coerce").fillna(0)
        style_df["Temporary Balance"] = np.select(
            [
                balance_values.isna().to_numpy(dtype=bool),
                (balance_values < 0).fillna(False).to_numpy(dtype=bool),
                (qty_values > 0).fillna(False).to_numpy(dtype=bool),
            ],
            [
                "background-color: #F3F4F6; color: #4B5563; font-weight: 850; text-align: right;",
                "background-color: #FDE2E1; color: #B42318; font-weight: 900; text-align: right;",
                "background-color: #DFF3E3; color: #067647; font-weight: 900; text-align: right;",
            ],
            default="background-color: #F3F4F6; color: #4B5563; font-weight: 850; text-align: right;",
        )

    styled_df = display_df.style.apply(lambda _: style_df, axis=None)
    numeric_subset = [col for col in ["Current Ending Balance", "Total Qty To Check", "Temporary Balance", "Shortage Qty"] if col in display_df.columns]
    if numeric_subset:
        styled_df = styled_df.set_properties(subset=numeric_subset, **{"text-align": "right"})
    st.dataframe(styled_df, use_container_width=True, hide_index=True, height=height)


def show_transaction_dataframe(df: pd.DataFrame, height: int = 420, limit: int = 500):
    total_rows = len(df)
    if total_rows > limit:
        st.caption(f"Showing first {limit:,} rows out of {total_rows:,} rows for faster loading.")
    else:
        st.caption(f"Showing {total_rows:,} rows.")

    display_df = df.head(limit).copy()

    if "Activity Date" in display_df.columns:
        display_df["Activity Date"] = pd.to_datetime(display_df["Activity Date"], errors="coerce").dt.strftime("%m/%d/%Y").replace("NaT", "")

    integer_cols = ["Excel Row", "Qty In", "Qty Out", "Balance After Transaction"]
    for col in integer_cols:
        if col in display_df.columns:
            display_df[col] = pd.to_numeric(display_df[col], errors="coerce").round(0).astype("Int64")

    for col in ["Is Not Shipped", "Is Cancelled"]:
        if col in display_df.columns:
            display_df[col] = display_df[col].fillna(False).astype(bool)

    numeric_subset = [col for col in ["Excel Row", "Qty In", "Qty Out", "Balance After Transaction"] if col in display_df.columns]
    center_subset = [col for col in ["Activity Date", "Transaction Type", "Is Not Shipped", "Is Cancelled"] if col in display_df.columns]
    left_subset = [col for col in ["Trans. #", "Ref #"] if col in display_df.columns]

    style_df = pd.DataFrame("", index=display_df.index, columns=display_df.columns)
    if "Transaction Type" in display_df.columns:
        transaction_types = display_df["Transaction Type"].astype(str).str.lower()
        has_inbound = transaction_types.str.contains("inbound", na=False)
        has_outbound = transaction_types.str.contains("outbound", na=False)
        style_df["Transaction Type"] = np.select(
            [has_inbound & ~has_outbound, has_outbound & ~has_inbound, has_inbound & has_outbound],
            [
                "background-color: #DFF3E3; color: #067647; font-weight: 700; text-align: center;",
                "background-color: #FDE2E1; color: #B42318; font-weight: 700; text-align: center;",
                "background-color: #E0F2FE; color: #026AA2; font-weight: 700; text-align: center;",
            ],
            default="background-color: #F3F4F6; color: #4B5563; font-weight: 700; text-align: center;",
        )

    styled_df = display_df.style.apply(lambda _: style_df, axis=None)
    if numeric_subset:
        styled_df = styled_df.set_properties(subset=numeric_subset, **{"text-align": "right"})
    if center_subset:
        styled_df = styled_df.set_properties(subset=center_subset, **{"text-align": "center"})
    if left_subset:
        styled_df = styled_df.set_properties(subset=left_subset, **{"text-align": "left"})

    column_config = {}
    if "Is Not Shipped" in display_df.columns:
        column_config["Is Not Shipped"] = st.column_config.CheckboxColumn(
            "Is Not Shipped",
            help="Checked when the transaction is marked Not Shipped.",
            disabled=True,
        )
    if "Is Cancelled" in display_df.columns:
        column_config["Is Cancelled"] = st.column_config.CheckboxColumn(
            "Is Cancelled",
            help="Checked when the transaction is marked Cancelled.",
            disabled=True,
        )

    st.dataframe(
        styled_df,
        use_container_width=True,
        hide_index=True,
        height=height,
        column_config=column_config,
    )


def reset_sidebar_filters(site_key):
    values = {
        f"{site_key}_filter_risk_levels": ["Data Issue", "Critical", "Warning", "Watch", "Healthy", "No Recent Demand"],
        f"{site_key}_filter_min_usage": 0,
        f"{site_key}_sku_select_combined": "",
    }
    for key, value in values.items():
        st.session_state[key] = value
    update_persistent_app_state(values=values)


def reset_transaction_filters(search_key, mode_key, date_key, range_key):
    st.session_state[search_key] = ""
    st.session_state[mode_key] = "All Dates"
    st.session_state.pop(date_key, None)
    st.session_state.pop(range_key, None)
    update_persistent_app_state(
        values={search_key: "", mode_key: "All Dates"},
        remove_keys=[date_key, range_key],
    )


restore_persistent_app_state()
st.sidebar.markdown(
    """
    <div class="sidebar-brand">
        <div class="fluent-grid-icon fluent-grid-icon-small" aria-hidden="true">
            <span></span><span></span><span></span><span></span>
        </div>
        <div class="sidebar-brand-copy">
            <div class="sidebar-brand-title">Inventory</div>
            <div class="sidebar-brand-subtitle">Operations Center</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)
if st.session_state.get("report_format") not in ["Newark", "Carson"]:
    st.session_state["report_format"] = "Newark"
st.sidebar.markdown('<div class="sidebar-section-title">Workspace</div>', unsafe_allow_html=True)
format_name = st.sidebar.selectbox("Warehouse", options=["Newark", "Carson"], index=0, key="report_format")
update_persistent_app_state(values={"report_format": format_name})
config = FORMAT_CONFIGS[format_name]
site_key = safe_format_slug(format_name).lower()
risk_filter_key = f"{site_key}_filter_risk_levels"
min_usage_filter_key = f"{site_key}_filter_min_usage"
sku_select_key = f"{site_key}_sku_select_combined"

st.sidebar.markdown('<div class="sidebar-section-title">Data source</div>', unsafe_allow_html=True)
st.sidebar.markdown('<div class="sidebar-section-help">Upload a report or continue from the last saved file.</div>', unsafe_allow_html=True)
uploaded = st.sidebar.file_uploader(
    config["upload_label"],
    type=["xlsx", "xls"],
    help=config["help"],
    label_visibility="collapsed",
    key=f"{format_name.lower()}_report_uploader",
)
report_source_slot = st.sidebar.empty()

st.sidebar.markdown('<div class="sidebar-section-title">Inventory filters</div>', unsafe_allow_html=True)
st.sidebar.markdown('<div class="sidebar-section-help">Filters apply to Overview and the SKU selector.</div>', unsafe_allow_html=True)
risk_options = ["Data Issue", "Critical", "Warning", "Watch", "Healthy", "No Recent Demand", "Inactive / No Demand"]
if risk_filter_key in st.session_state:
    st.session_state[risk_filter_key] = [value for value in st.session_state[risk_filter_key] if value in risk_options]
show_risks = st.sidebar.multiselect(
    "Risk Level",
    options=risk_options,
    default=["Data Issue", "Critical", "Warning", "Watch", "Healthy", "No Recent Demand"],
    key=risk_filter_key,
)
min_usage = st.sidebar.number_input(
    "Min 30D Outbound",
    min_value=0,
    value=0,
    step=1,
    key=min_usage_filter_key,
)
update_persistent_app_state(values={risk_filter_key: show_risks, min_usage_filter_key: min_usage})

sku_sidebar_slot = st.sidebar.empty()
st.sidebar.button("Clear filters", use_container_width=True, on_click=reset_sidebar_filters, args=(site_key,))


status_box = st.empty()

using_saved_report = False
active_file_name = ""
saved_upload_cache_key = f"_saved_upload_cache_{site_key}"
uploaded_hash_cache_key = f"_uploaded_hash_cache_{site_key}"
saved_upload_identity_key = f"_saved_upload_identity_{site_key}"

if uploaded is not None:
    file_bytes = uploaded.getvalue()
    active_file_name = uploaded.name
    uploaded_file_id = clean_text(getattr(uploaded, "file_id", ""))
    uploaded_hash_cache = st.session_state.get(uploaded_hash_cache_key, {})
    if (
        uploaded_file_id
        and isinstance(uploaded_hash_cache, dict)
        and uploaded_hash_cache.get("file_id") == uploaded_file_id
        and uploaded_hash_cache.get("file_name") == active_file_name
        and uploaded_hash_cache.get("size") == len(file_bytes)
    ):
        file_hash = uploaded_hash_cache.get("sha256", "")
    else:
        file_hash = stable_file_hash(file_bytes)
        st.session_state[uploaded_hash_cache_key] = {
            "file_id": uploaded_file_id,
            "file_name": active_file_name,
            "size": len(file_bytes),
            "sha256": file_hash,
        }
    if not file_hash:
        file_hash = stable_file_hash(file_bytes)
        st.session_state[uploaded_hash_cache_key] = {
            "file_id": uploaded_file_id,
            "file_name": active_file_name,
            "size": len(file_bytes),
            "sha256": file_hash,
        }
    upload_identity = f"{active_file_name}|{len(file_bytes)}|{file_hash}"
    if st.session_state.get(saved_upload_identity_key) != upload_identity:
        save_persistent_upload(format_name, active_file_name, file_bytes, file_hash)
        st.session_state[saved_upload_identity_key] = upload_identity
    st.session_state[saved_upload_cache_key] = {
        "file_bytes": file_bytes,
        "file_name": active_file_name,
        "size": len(file_bytes),
        "sha256": file_hash,
        "last_selected_format": format_name,
    }
else:
    saved_upload = st.session_state.get(saved_upload_cache_key)
    if not isinstance(saved_upload, dict):
        saved_upload = load_persistent_upload(format_name)
        if saved_upload is not None:
            st.session_state[saved_upload_cache_key] = saved_upload
    if saved_upload is None:
        st.markdown(
            f"""
            <div class="empty-state">
                <div class="empty-state-title">Upload a {html.escape(format_name)} report</div>
                <div class="empty-state-text">Use the report uploader in the sidebar. After the first successful upload, the dashboard will automatically reopen the saved report.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        report_source_slot.markdown(
            '<div class="data-source-card data-source-card-empty"><span>No saved report</span></div>',
            unsafe_allow_html=True,
        )
        st.stop()
    file_bytes = saved_upload["file_bytes"]
    active_file_name = saved_upload["file_name"]
    file_hash = saved_upload.get("sha256") or stable_file_hash(file_bytes)
    using_saved_report = True

with report_source_slot.container():
    report_source_col, report_refresh_col = st.columns([5, 1], gap="small")
    with report_source_col:
        st.markdown(
            f'<div class="data-source-card"><span>{html.escape("Saved" if using_saved_report else "Uploaded")}: {html.escape(active_file_name)}</span></div>',
            unsafe_allow_html=True,
        )
    with report_refresh_col:
        refresh_report_clicked = st.button(
            "↻",
            help="Refresh full report",
            key=f"refresh_full_report_{site_key}",
            use_container_width=True,
        )

report_refresh_version_key = f"_report_refresh_version_{site_key}"
if refresh_report_clicked:
    st.session_state[report_refresh_version_key] = int(st.session_state.get(report_refresh_version_key, 0)) + 1
    refresh_keys = [
        saved_upload_cache_key,
        uploaded_hash_cache_key,
        saved_upload_identity_key,
        f"_inventory_model_{site_key}",
        f"_inventory_model_source_{site_key}",
        f"_do_search_text_{site_key}",
        f"_stock_do_tables_{site_key}",
        "last_upload_effect_key",
    ]
    for key in refresh_keys:
        st.session_state.pop(key, None)
    st.rerun()

uploaded_key = f"{format_name}|{active_file_name}|{len(file_bytes)}|{file_hash}"
show_upload_effect = uploaded is not None and st.session_state.get("last_upload_effect_key") != uploaded_key
model_cache_key = f"_inventory_model_{site_key}"
model_source_key = f"_inventory_model_source_{site_key}"
report_refresh_version = int(st.session_state.get(report_refresh_version_key, 0))
model_cache_version = f"{APP_CACHE_VERSION}|refresh-{report_refresh_version}"
model_source_value = f"{uploaded_key}|{model_cache_version}"

try:
    if show_upload_effect:
        status_box.markdown(
            f"""
            <div class="loading-stage-card">
                <div class="loading-row">
                    <div class="loader-ring"></div>
                    <div class="stage-copy">
                        <div class="stage-title">Loading Item Activity Report</div>
                        <div class="stage-subtitle">Reading the workbook, validating the {html.escape(format_name)} format, and building the inventory model.</div>
                        <div class="stage-meta">{html.escape(active_file_name)}</div>
                    </div>
                </div>
                <div class="animated-progress"></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    if st.session_state.get(model_source_key) == model_source_value and isinstance(st.session_state.get(model_cache_key), dict):
        model = st.session_state[model_cache_key]
    else:
        model = process_excel_file(file_bytes, format_name, model_cache_version)
        st.session_state[model_cache_key] = model
        st.session_state[model_source_key] = model_source_value

    if not show_upload_effect:
        status_box.empty()
except WrongFileFormatError as exc:
    status_box.markdown(
        f"""
        <div class="error-stage-card">
            <div class="error-row">
                <div class="error-mark">!</div>
                <div class="stage-copy">
                    <div class="stage-title">Report format not recognized</div>
                    <div class="stage-subtitle">{html.escape(str(exc))}</div>
                    <div class="stage-meta">Confirm the warehouse format and upload the matching Item Activity Report.</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()
except Exception:
    status_box.markdown(
        """
        <div class="error-stage-card">
            <div class="error-row">
                <div class="error-mark">!</div>
                <div class="stage-copy">
                    <div class="stage-title">File could not be processed</div>
                    <div class="stage-subtitle">Check the selected warehouse format and upload a valid Item Activity Report.</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()

sku_df = model["sku_df"]

sku_option_source = sku_df
if show_risks:
    sku_option_source = sku_option_source[sku_option_source["Risk Level"].isin(show_risks)]
sku_option_source = sku_option_source[sku_option_source["Outbound Last 30 Days"] >= min_usage]
sku_options = [""] + sku_option_source["SKU"].astype(str).dropna().tolist()

with sku_sidebar_slot.container():
    st.markdown('<div class="sidebar-section-gap"></div>', unsafe_allow_html=True)

    if st.session_state.get(sku_select_key) not in sku_options:
        st.session_state[sku_select_key] = ""
        update_persistent_app_state(values={sku_select_key: ""})

    selected_sku = st.selectbox(
        "Search / Select SKU",
        options=sku_options,
        index=0,
        key=sku_select_key,
        format_func=lambda x: "" if x == "" else x,
        help="Select a SKU to show only that SKU. Leave blank to use Risk Level filters.",
    )

selected_sku = clean_text(selected_sku)
update_persistent_app_state(values={sku_select_key: selected_sku})

if selected_sku:
    priority_filtered = sku_df[sku_df["SKU"].astype(str) == str(selected_sku)].copy()
else:
    priority_filtered = sku_option_source.copy()

report_start = model["report_start"]
report_end = model["report_end"]
windows = model["windows"]

source_label = "Saved" if using_saved_report else "Uploaded"
st.markdown(
    f"""
    <div class="app-header">
        <div class="app-header-main">
            <div class="app-title-cluster">
                <div class="fluent-grid-icon app-product-icon" aria-hidden="true">
                    <span></span><span></span><span></span><span></span>
                </div>
                <div class="app-title-copy">
                    <div class="app-eyebrow">{html.escape(format_name)} warehouse</div>
                    <div class="app-title">Inventory Shortage</div>
                </div>
            </div>
            <div class="app-subtitle">Inventory risk, SKU activity, DO lookup, and outbound stock validation.</div>
        </div>
        <div class="app-meta">
            <span class="meta-chip meta-chip-date meta-chip-accent">{fmt_date(report_start)} – {fmt_date(report_end)}</span>
            <span class="meta-chip meta-chip-file" title="{html.escape(active_file_name)}">{html.escape(active_file_name)}</span>
            <span class="meta-chip meta-chip-status">{source_label}</span>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

navigation_options = ["Overview", "SKU Detail", "DO Lookup", "Stock Check", "Audit", "Help"]
if st.session_state.get("main_page_navigation") not in navigation_options:
    st.session_state["main_page_navigation"] = "Overview"
with st.container(key="main_navigation"):
    if hasattr(st, "segmented_control"):
        selected_page = st.segmented_control(
            "Navigation",
            options=navigation_options,
            key="main_page_navigation",
            selection_mode="single",
            label_visibility="collapsed",
        )
    else:
        selected_page = st.radio(
            "Navigation",
            options=navigation_options,
            key="main_page_navigation",
            horizontal=True,
            label_visibility="collapsed",
        )
selected_page = selected_page or "Overview"
update_persistent_app_state(values={"main_page_navigation": selected_page})

if selected_page == "Overview":
    data_issue_count = int((sku_df["Risk Level"] == "Data Issue").sum())
    critical_count = int((sku_df["Risk Level"] == "Critical").sum())
    warning_count = int((sku_df["Risk Level"] == "Warning").sum())
    watch_count = int((sku_df["Risk Level"] == "Watch").sum())
    healthy_count = int((sku_df["Risk Level"] == "Healthy").sum())
    no_demand_count = int((sku_df["Risk Level"] == "No Recent Demand").sum())
    inactive_count = int((sku_df["Risk Level"] == "Inactive / No Demand").sum())
    active_count = int((sku_df["Demand Status"] == "Active").sum())
    dormant_count = int((sku_df["Demand Status"] == "Dormant").sum())

    export_file_name = report_download_filename(format_name, report_end)
    transaction_file_name = transaction_download_filename(format_name, report_end)
    report_export_key = f"report_export_{uploaded_key}"
    transaction_export_key = f"transaction_export_{uploaded_key}"
    report_export_data = lambda export_model=model, export_format=format_name: to_excel_bytes(export_model, export_format, CUSTOMER_EXPORT_VERSION)
    transaction_export_data = lambda export_model=model, export_format=format_name: to_transaction_excel_bytes(export_model, export_format, APP_CACHE_VERSION)
    heading_col, export_col_1, export_col_2 = st.columns([3.4, 1.25, 1.25])
    with heading_col:
        tab_page_header("Overview", "Review inventory risk first, then move into SKU, DO, or stock validation workflows.")
    with export_col_1:
        st.download_button(
            "Download report",
            data=report_export_data,
            file_name=export_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            help=f"Download {export_file_name}",
            key=f"download_{report_export_key}",
            on_click="ignore",
        )
    with export_col_2:
        st.download_button(
            "Download transactions",
            data=transaction_export_data,
            file_name=transaction_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help=f"Download {transaction_file_name}",
            key=f"download_{transaction_export_key}",
            on_click="ignore",
        )

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        metric_card("Total SKUs", fmt_num(len(sku_df)), f"Active {active_count:,} · Dormant {dormant_count:,} · Inactive {inactive_count:,}")
    with k2:
        metric_card("Critical SKUs", fmt_num(critical_count), f"Data issues {data_issue_count:,}")
    with k3:
        metric_card("Warning SKUs", fmt_num(warning_count), "Need ETA or reserve review")
    with k4:
        metric_card("Watch SKUs", fmt_num(watch_count), "Monitor usage trend")

    st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Shortage Priority List</div>", unsafe_allow_html=True)
    st.markdown(
        f"<div class='section-subtitle'>Showing {len(priority_filtered):,} of {len(sku_df):,} SKUs. Critical applies only to active 30-day demand; inactive SKUs are hidden by default.</div>",
        unsafe_allow_html=True,
    )

    priority_cols = [
        "SKU",
        "Description",
        "Risk Level",
        "Ending Balance",
        "Last Outbound Date",
        "Avg Daily Usage 30D",
        "Days Remaining",
        "Forecast Stockout Date",
    ]
    priority_display = prepare_display(priority_filtered[priority_cols])
    priority_height = min(620, max(300, 88 + (min(len(priority_display), 16) * 31)))
    show_limited_dataframe(priority_display, height=priority_height, limit=250, show_count=False)


elif selected_page == "SKU Detail":
    tab_page_header("SKU Detail", "Review one SKU, its risk position, source metrics, and complete transaction history.")
    if not selected_sku:
        st.info("Select a SKU to view SKU Detail.")
    elif priority_filtered.empty:
        st.warning("No SKU matches the current filters.")
    else:
        selected = sku_df[sku_df["SKU"].astype(str) == str(selected_sku)].iloc[0]
        selected_sku_safe = html.escape(str(selected_sku))
        selected_description_safe = html.escape(clean_text(selected["Description"]))
        st.markdown(
            f"""
            <div class="selected-sku-card">
                <div class="selected-sku-label">Selected SKU</div>
                <div class="selected-sku-value">{selected_sku_safe}</div>
                <div class="selected-sku-description">{selected_description_safe}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        ending_balance_value = pd.to_numeric(pd.Series([selected["Ending Balance"]]), errors="coerce").iloc[0]
        if pd.isna(ending_balance_value):
            st.warning("This SKU is missing a valid Ending Balance value.")
        elif ending_balance_value <= 0 and selected["Risk Level"] == "Critical":
            st.warning("This SKU has active demand and zero or negative ending balance.")
        elif selected["Risk Level"] == "No Recent Demand":
            st.info("This SKU had outbound demand in the 90-day data window but none in the 30-day data window.")
        elif selected["Risk Level"] == "Inactive / No Demand":
            st.info("This SKU has no outbound demand in the 90-day data window.")

        d1, d2, d3, d4 = st.columns(4)
        with d1:
            metric_card("Risk Level", risk_badge_text(selected["Risk Level"]), selected["Recommended Action"])
        with d2:
            metric_card("Ending Balance", fmt_num(selected["Ending Balance"]), "Official ending balance")
        with d3:
            metric_card("Days Remaining", fmt_num(selected["Days Remaining"], 1), "Based on Avg Daily Usage 30D")
        with d4:
            metric_card("Forecast Stockout", fmt_date(selected["Forecast Stockout Date"]), "Business day estimate")

        st.markdown("<div class='section-block'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title">SKU metrics</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-subtitle">Demand, usage, and recent activity for the selected item.</div>', unsafe_allow_html=True)
        detail_cols = [
            "Demand Status",
            "Official Total Inbound",
            "Official Total Outbound",
            "Avg Daily Usage 30D",
            "Last Outbound Date",
            "Last Inbound Date",
            "Last Activity Date",
            "Official Total Row",
            "Official Ending Row",
        ]
        detail = selected[detail_cols].to_frame("Value")
        detail["Value"] = detail.apply(
            lambda r: fmt_date(r["Value"]) if "date" in str(r.name).lower()
            else (
                fmt_num(r["Value"], 2) if str(r.name) == "Avg Daily Usage 30D"
                else (fmt_num(r["Value"]) if isinstance(r["Value"], (int, float, np.integer, np.floating)) and np.isfinite(r["Value"]) else r["Value"])
            ),
            axis=1,
        )
        st.dataframe(detail, use_container_width=True)

        tx_source = model["tx_df"]
        if not tx_source.empty:
            tx_sku = tx_source.loc[tx_source["SKU"] == selected_sku].copy()
            st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
            st.markdown('<div class="section-title">Transaction history</div>', unsafe_allow_html=True)
            st.markdown('<div class="section-subtitle">Search references or filter activity dates without leaving this SKU.</div>', unsafe_allow_html=True)
            full_tx_cols = [
                "Excel Row",
                "Activity Date",
                "Transaction Type",
                "Trans. #",
                "Ref #",
                "Qty In",
                "Qty Out",
                "Balance After Transaction",
                "Is Not Shipped",
                "Is Cancelled",
            ]
            for col in full_tx_cols:
                if col not in tx_sku.columns:
                    if col in ["Qty In", "Qty Out", "Balance After Transaction"]:
                        tx_sku[col] = 0.0
                    elif col in ["Is Not Shipped", "Is Cancelled"]:
                        tx_sku[col] = False
                    elif col == "Activity Date":
                        tx_sku[col] = pd.NaT
                    else:
                        tx_sku[col] = ""

            if tx_sku.empty:
                st.info("No transaction history found for this SKU.")
            else:
                total_qty_in = pd.to_numeric(tx_sku["Qty In"], errors="coerce").fillna(0).sum()
                total_qty_out = pd.to_numeric(tx_sku["Qty Out"], errors="coerce").fillna(0).sum()

                s1, s2, s3 = st.columns(3)
                with s1:
                    metric_card("Total Qty In", fmt_num(total_qty_in), "Inbound transaction total")
                with s2:
                    metric_card("Total Qty Out", fmt_num(total_qty_out), "Outbound transaction total")
                with s3:
                    metric_card("Current Balance", fmt_num(selected["Ending Balance"]), "Official ending balance")

                st.markdown("<div class='kpi-row-gap'></div>", unsafe_allow_html=True)

                sku_filter_key = re.sub(r"[^A-Za-z0-9_]+", "_", f"{format_name}_{selected_sku}")[:55]
                tx_filtered = tx_sku
                tx_search_key = f"tx_search_{sku_filter_key}"
                tx_mode_key = f"tx_date_mode_{sku_filter_key}"
                tx_date_key = f"tx_date_{sku_filter_key}"
                tx_range_key = f"tx_date_range_{sku_filter_key}"

                tx_dates = pd.to_datetime(tx_sku["Activity Date"], errors="coerce").dropna()
                if not tx_dates.empty:
                    tx_min_date = tx_dates.min().date()
                    tx_max_date = tx_dates.max().date()
                else:
                    tx_min_date = None
                    tx_max_date = None

                st.markdown(
                    """
                    <div class="tx-filter-shell">
                        <div class="tx-filter-top">
                            <div>
                                <div class="tx-filter-title">Transaction Search Workspace</div>
                                <div class="tx-filter-subtitle">Paste multiple Ref # / Trans. # values, select a date mode, and review matched or missing values before checking the table.</div>
                            </div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                header_left, header_right = st.columns([5, 1.15])
                with header_left:
                    st.markdown("<div class='section-title'>Transaction Filters</div>", unsafe_allow_html=True)
                    st.markdown("<div class='section-subtitle'>Use line breaks, commas, or semicolons for multiple search values.</div>", unsafe_allow_html=True)
                with header_right:
                    st.button(
                        "Clear Filters",
                        use_container_width=True,
                        key=f"clear_tx_filters_{sku_filter_key}",
                        on_click=reset_transaction_filters,
                        args=(tx_search_key, tx_mode_key, tx_date_key, tx_range_key),
                    )

                search_col, date_col = st.columns([1.55, 1], gap="medium")
                with search_col:
                    st.markdown(
                        """
                        <div class="tx-filter-card-title">Search Ref # / Trans. #</div>
                        <div class="tx-filter-card-subtitle">Paste one value per line for the cleanest result. Matching is not case-sensitive.</div>
                        <div class="tx-example-row">
                            <span class="tx-example-pill">PO_0090</span>
                            <span class="tx-example-pill">PO_0086</span>
                            <span class="tx-example-pill">AXIA_2484</span>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                    tx_search = st.text_area(
                        "Search Ref # / Trans. #",
                        placeholder="PO_0090\nPO_0086\nAXIA_2484",
                        key=tx_search_key,
                        height=154,
                        label_visibility="collapsed",
                    )
                    update_persistent_app_state(values={tx_search_key: tx_search})
                with date_col:
                    st.markdown(
                        """
                        <div class="tx-filter-card-title">Activity Date Filter</div>
                        <div class="tx-filter-card-subtitle">Choose all dates, one exact date, or a date range.</div>
                        """,
                        unsafe_allow_html=True,
                    )
                    if st.session_state.get(tx_mode_key) not in ["All Dates", "Single Date", "Date Range"]:
                        st.session_state[tx_mode_key] = "All Dates"
                    activity_date_mode = st.radio(
                        "Activity Date Filter",
                        options=["All Dates", "Single Date", "Date Range"],
                        horizontal=True,
                        key=tx_mode_key,
                        label_visibility="collapsed",
                    )
                    update_persistent_app_state(values={tx_mode_key: activity_date_mode})
                    selected_tx_date = None
                    selected_tx_date_range = None
                    if activity_date_mode == "Single Date":
                        saved_tx_date = st.session_state.get(tx_date_key)
                        if saved_tx_date is not None:
                            saved_tx_date = pd.to_datetime(saved_tx_date, errors="coerce")
                            if pd.isna(saved_tx_date) or (tx_min_date is not None and saved_tx_date.date() < tx_min_date) or (tx_max_date is not None and saved_tx_date.date() > tx_max_date):
                                st.session_state.pop(tx_date_key, None)
                                update_persistent_app_state(remove_keys=[tx_date_key])
                        selected_tx_date = st.date_input(
                            "Select Activity Date",
                            value=None,
                            min_value=tx_min_date,
                            max_value=tx_max_date,
                            key=tx_date_key,
                        )
                        update_persistent_app_state(values={tx_date_key: selected_tx_date})
                    elif activity_date_mode == "Date Range":
                        default_tx_date_range = (tx_min_date, tx_max_date) if tx_min_date is not None and tx_max_date is not None else None
                        saved_tx_range = st.session_state.get(tx_range_key)
                        if isinstance(saved_tx_range, (list, tuple)) and len(saved_tx_range) == 2:
                            saved_start = pd.to_datetime(saved_tx_range[0], errors="coerce")
                            saved_end = pd.to_datetime(saved_tx_range[1], errors="coerce")
                            if pd.isna(saved_start) or pd.isna(saved_end) or (tx_min_date is not None and saved_start.date() < tx_min_date) or (tx_max_date is not None and saved_end.date() > tx_max_date):
                                st.session_state.pop(tx_range_key, None)
                                update_persistent_app_state(remove_keys=[tx_range_key])
                        selected_tx_date_range = st.date_input(
                            "Select Activity Date Range",
                            value=default_tx_date_range,
                            min_value=tx_min_date,
                            max_value=tx_max_date,
                            key=tx_range_key,
                        )
                        update_persistent_app_state(values={tx_range_key: selected_tx_date_range})
                    else:
                        if tx_min_date is not None and tx_max_date is not None:
                            st.markdown(
                                f"""
                                <div class="tx-result-box">
                                    <div class="tx-result-title">Available Activity Dates</div>
                                    <div class="tx-pill-wrap">
                                        <span class="tx-pill-muted">{tx_min_date.strftime('%m/%d/%Y')}</span>
                                        <span class="tx-pill-muted">to</span>
                                        <span class="tx-pill-muted">{tx_max_date.strftime('%m/%d/%Y')}</span>
                                    </div>
                                </div>
                                """,
                                unsafe_allow_html=True,
                            )

                tx_terms = []
                if tx_search.strip():
                    tx_terms = [x.strip() for x in re.split(r"[\n,;]+", tx_search) if x.strip()]
                    tx_pattern = "|".join(re.escape(x.lower()) for x in tx_terms)
                    tx_filtered = tx_filtered[
                        tx_filtered["Ref #"].astype(str).str.lower().str.contains(tx_pattern, na=False, regex=True)
                        | tx_filtered["Trans. #"].astype(str).str.lower().str.contains(tx_pattern, na=False, regex=True)
                    ]

                date_filter_label = "All Dates"
                date_filter_help = "No date filter"
                if activity_date_mode == "Single Date" and selected_tx_date is not None:
                    selected_date_value = pd.to_datetime(selected_tx_date).normalize()
                    date_filter_label = selected_date_value.strftime("%m/%d/%Y")
                    date_filter_help = "Single date"
                    tx_activity_dates = pd.to_datetime(tx_filtered["Activity Date"], errors="coerce").dt.normalize()
                    tx_filtered = tx_filtered[tx_activity_dates == selected_date_value]
                elif activity_date_mode == "Date Range" and selected_tx_date_range is not None:
                    if isinstance(selected_tx_date_range, tuple) and len(selected_tx_date_range) == 2:
                        range_start = pd.to_datetime(selected_tx_date_range[0]).normalize()
                        range_end = pd.to_datetime(selected_tx_date_range[1]).normalize()
                        if range_start > range_end:
                            range_start, range_end = range_end, range_start
                        date_filter_label = f"{range_start.strftime('%m/%d/%Y')} - {range_end.strftime('%m/%d/%Y')}"
                        date_filter_help = "Date range"
                        tx_activity_dates = pd.to_datetime(tx_filtered["Activity Date"], errors="coerce").dt.normalize()
                        tx_filtered = tx_filtered[(tx_activity_dates >= range_start) & (tx_activity_dates <= range_end)]
                    else:
                        st.warning("Please select both start and end date for Activity Date Range.")
                        date_filter_label = "Range incomplete"
                        date_filter_help = "Select 2 dates"

                tx_missing_terms = []
                tx_found_terms = []
                if tx_terms:
                    tx_result_text = (
                        tx_filtered["Ref #"].astype(str).str.lower()
                        + " "
                        + tx_filtered["Trans. #"].astype(str).str.lower()
                    )
                    tx_missing_terms = [
                        term for term in tx_terms
                        if not tx_result_text.str.contains(re.escape(term.lower()), na=False, regex=True).any()
                    ]
                    tx_found_terms = [term for term in tx_terms if term not in tx_missing_terms]

                search_value_count = len(tx_terms)
                missing_value_count = len(tx_missing_terms)
                matching_row_count = len(tx_filtered)
                found_value_count = search_value_count - missing_value_count
                search_status_text = f"{found_value_count:,} / {search_value_count:,}" if search_value_count else "0"
                search_status_help = "found searched values" if search_value_count else "no search values pasted"
                missing_status_help = "needs review" if missing_value_count else "all clear"

                st.markdown(
                    f"""
                    <div class="tx-status-grid">
                        <div class="tx-status-card">
                            <div class="tx-status-label">Search Values</div>
                            <div class="tx-status-value">{search_value_count:,}</div>
                            <div class="tx-status-help">Ref # / Trans. # entered</div>
                        </div>
                        <div class="tx-status-card">
                            <div class="tx-status-label">Matched Values</div>
                            <div class="tx-status-value">{html.escape(search_status_text)}</div>
                            <div class="tx-status-help">{html.escape(search_status_help)}</div>
                        </div>
                        <div class="tx-status-card">
                            <div class="tx-status-label">Matching Rows</div>
                            <div class="tx-status-value">{matching_row_count:,}</div>
                            <div class="tx-status-help">after all filters</div>
                        </div>
                        <div class="tx-status-card">
                            <div class="tx-status-label">Date Filter</div>
                            <div class="tx-status-value" style="font-size:.98rem; line-height:1.2; word-break:break-word;">{html.escape(date_filter_label)}</div>
                            <div class="tx-status-help">{html.escape(date_filter_help)}</div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

                if tx_terms:
                    if tx_missing_terms:
                        st.markdown(
                            f"""
                            <div class="tx-result-box tx-result-box-missing">
                                <div class="tx-result-title">Not found in filtered results</div>
                                <div class="tx-pill-wrap">{html_pills(tx_missing_terms, 'tx-pill-missing')}</div>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            """
                            <div class="tx-result-box tx-result-box-ok">
                                <div class="tx-result-title">All searched values were found</div>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )
                else:
                    st.markdown(
                        f"""
                        <div class="tx-result-box">
                            <div class="tx-result-title">Current view</div>
                            <div class="tx-pill-wrap"><span class="tx-pill-muted">Showing {matching_row_count:,} transaction rows for the selected SKU and date filter</span></div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

                st.markdown("<div class='kpi-row-gap'></div>", unsafe_allow_html=True)

                tx_filtered = tx_filtered.sort_values(["Activity Date", "Excel Row"], ascending=[False, False])

                show_transaction_dataframe(tx_filtered[full_tx_cols], height=420, limit=500)


elif selected_page == "DO Lookup":
    tab_page_header("DO Lookup", "Paste one or multiple DO numbers to find matching items across every SKU and review results by DO.")

    do_tx = model["tx_df"]
    do_lookup_key = f"do_lookup_{site_key}"
    do_clear_key = f"clear_do_lookup_{site_key}"

    do_header_left, do_header_right = st.columns([5, 1.15])
    with do_header_left:
        st.markdown("<div class='section-title compact-heading'>Search DO #</div>", unsafe_allow_html=True)
        st.markdown("<div class='section-subtitle'>Use line breaks, commas, or semicolons. Each DO will display in its own grouped section.</div>", unsafe_allow_html=True)
    with do_header_right:
        if st.button("Clear", use_container_width=True, key=do_clear_key):
            st.session_state[do_lookup_key] = ""
            update_persistent_app_state(values={do_lookup_key: ""})
            st.rerun()

    do_lookup_value = st.text_area(
        "Search DO #",
        placeholder="AXIA_2484\nPO_0090",
        key=do_lookup_key,
        height=104,
        label_visibility="collapsed",
    )
    update_persistent_app_state(values={do_lookup_key: do_lookup_value})

    do_lookup_terms = []
    seen_do_terms = set()
    for value in re.split(r"[\n,;]+", do_lookup_value):
        term = value.strip()
        term_key = term.lower()
        if term and term_key not in seen_do_terms:
            do_lookup_terms.append(term)
            seen_do_terms.add(term_key)

    if do_tx.empty:
        st.info("No transaction data found in this report.")
    elif not do_lookup_terms:
        st.markdown(
            """
            <div class="tx-result-box">
                <div class="tx-result-title">Waiting for DO #</div>
                <div class="tx-pill-wrap"><span class="tx-pill-muted">Enter a DO # to show all items belonging to it.</span></div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        do_required_cols = ["Ref #", "Trans. #", "SKU", "Description", "Activity Date", "Transaction Type", "Qty In", "Qty Out", "Balance After Transaction", "Is Not Shipped", "Is Cancelled", "Excel Row"]
        if any(col not in do_tx.columns for col in do_required_cols):
            do_tx = do_tx.copy()
        for col in do_required_cols:
            if col not in do_tx.columns:
                if col in ["Qty In", "Qty Out", "Balance After Transaction"]:
                    do_tx[col] = 0.0
                elif col in ["Is Not Shipped", "Is Cancelled"]:
                    do_tx[col] = False
                elif col == "Activity Date":
                    do_tx[col] = pd.NaT
                else:
                    do_tx[col] = ""

        do_search_cache_key = f"_do_search_text_{site_key}"
        do_search_cache = st.session_state.get(do_search_cache_key, {})
        if isinstance(do_search_cache, dict) and do_search_cache.get("source") == model_source_value and isinstance(do_search_cache.get("values"), pd.Series):
            do_search_text = do_search_cache["values"]
        else:
            do_search_text = do_tx["Ref #"].astype(str).str.lower() + " " + do_tx["Trans. #"].astype(str).str.lower()
            st.session_state[do_search_cache_key] = {"source": model_source_value, "values": do_search_text}

        do_found_terms = []
        do_missing_terms = []
        do_matched_frames = []

        for term in do_lookup_terms:
            term_mask = do_search_text.str.contains(re.escape(term.lower()), na=False, regex=True)
            if term_mask.any():
                do_found_terms.append(term)
                term_df = do_tx[term_mask].copy()
                term_df.insert(0, "Searched DO #", term)
                do_matched_frames.append(term_df)
            else:
                do_missing_terms.append(term)

        do_detail_df = pd.concat(do_matched_frames, ignore_index=True) if do_matched_frames else pd.DataFrame(columns=["Searched DO #"] + list(do_tx.columns))

        unique_sku_count = do_detail_df["SKU"].astype(str).replace("", np.nan).dropna().nunique() if not do_detail_df.empty else 0
        total_do_qty_out = pd.to_numeric(do_detail_df["Qty Out"], errors="coerce").fillna(0).sum() if not do_detail_df.empty else 0
        total_do_qty_in = pd.to_numeric(do_detail_df["Qty In"], errors="coerce").fillna(0).sum() if not do_detail_df.empty else 0
        matched_value_text = f"{len(do_found_terms):,} / {len(do_lookup_terms):,}"

        st.markdown(
            f"""
            <div class="tx-status-grid">
                <div class="tx-status-card">
                    <div class="tx-status-label">Search Values</div>
                    <div class="tx-status-value">{len(do_lookup_terms):,}</div>
                    <div class="tx-status-help">DO # entered</div>
                </div>
                <div class="tx-status-card">
                    <div class="tx-status-label">Matched Values</div>
                    <div class="tx-status-value">{html.escape(matched_value_text)}</div>
                    <div class="tx-status-help">found in DO # / Trans. #</div>
                </div>
                <div class="tx-status-card">
                    <div class="tx-status-label">Items / SKUs</div>
                    <div class="tx-status-value">{unique_sku_count:,}</div>
                    <div class="tx-status-help">unique SKU count</div>
                </div>
                <div class="tx-status-card">
                    <div class="tx-status-label">Qty In / Out</div>
                    <div class="tx-status-value">{fmt_num(total_do_qty_in)} / {fmt_num(total_do_qty_out)}</div>
                    <div class="tx-status-help">matched rows total</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if do_missing_terms:
            st.markdown(
                f"""
                <div class="tx-result-box tx-result-box-missing">
                    <div class="tx-result-title">Not found in full transaction history</div>
                    <div class="tx-pill-wrap">{html_pills(do_missing_terms, 'tx-pill-missing')}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        if do_found_terms:
            st.markdown(
                f"""
                <div class="tx-result-box tx-result-box-ok">
                    <div class="tx-result-title">Found DO #</div>
                    <div class="tx-pill-wrap">{html_pills(do_found_terms, 'tx-pill-ok')}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

        if do_detail_df.empty:
            st.info("No matching rows found for the entered DO #.")
        else:
            do_detail_df["Activity Date"] = pd.to_datetime(do_detail_df["Activity Date"], errors="coerce")
            do_detail_df["Qty In"] = pd.to_numeric(do_detail_df["Qty In"], errors="coerce").fillna(0)
            do_detail_df["Qty Out"] = pd.to_numeric(do_detail_df["Qty Out"], errors="coerce").fillna(0)

            overview_rows = []
            for term in do_lookup_terms:
                term_detail_df = do_detail_df[do_detail_df["Searched DO #"] == term].copy()
                if term_detail_df.empty:
                    overview_rows.append(
                        {
                            "DO #": term,
                            "Status": "Not Found",
                            "SKU Count": 0,
                            "Qty In": 0,
                            "Qty Out": 0,
                            "Activity Date Range": "-",
                        }
                    )
                else:
                    term_dates = pd.to_datetime(term_detail_df["Activity Date"], errors="coerce").dropna()
                    if term_dates.empty:
                        date_range_text = "-"
                    else:
                        date_start = term_dates.min().strftime("%m/%d/%Y")
                        date_end = term_dates.max().strftime("%m/%d/%Y")
                        date_range_text = date_start if date_start == date_end else f"{date_start} - {date_end}"
                    overview_rows.append(
                        {
                            "DO #": term,
                            "Status": "Found",
                            "SKU Count": term_detail_df["SKU"].astype(str).replace("", np.nan).dropna().nunique(),
                            "Qty In": pd.to_numeric(term_detail_df["Qty In"], errors="coerce").fillna(0).sum(),
                            "Qty Out": pd.to_numeric(term_detail_df["Qty Out"], errors="coerce").fillna(0).sum(),
                            "Activity Date Range": date_range_text,
                        }
                    )

            st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>DO Search Overview</div>", unsafe_allow_html=True)
            st.markdown("<div class='section-subtitle'>Shows each pasted DO # in the same order entered, so missing and matched values are easy to review.</div>", unsafe_allow_html=True)
            overview_df = pd.DataFrame(overview_rows)
            overview_height = min(310, max(132, 74 + (len(overview_df) * 32)))
            show_limited_dataframe(overview_df, height=overview_height, limit=1000, show_count=False)

            st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>Items Belonging to Each DO #</div>", unsafe_allow_html=True)
            st.markdown("<div class='section-subtitle'>Each section below is separated by the exact DO # you searched.</div>", unsafe_allow_html=True)

            for term in do_found_terms:
                term_detail_df = do_detail_df[do_detail_df["Searched DO #"] == term].copy()
                term_summary = (
                    term_detail_df.groupby(["Searched DO #", "SKU", "Description"], dropna=False)
                    .agg(
                        **{
                            "First Activity Date": ("Activity Date", "min"),
                            "Latest Activity Date": ("Activity Date", "max"),
                            "Total Qty In": ("Qty In", "sum"),
                            "Total Qty Out": ("Qty Out", "sum"),
                        }
                    )
                    .reset_index()
                )
                term_summary = term_summary.sort_values(["SKU"], ascending=[True]).reset_index(drop=True)
                term_summary = term_summary.rename(columns={"Searched DO #": "DO #"})
                term_qty_out = pd.to_numeric(term_detail_df["Qty Out"], errors="coerce").fillna(0).sum()
                term_qty_in = pd.to_numeric(term_detail_df["Qty In"], errors="coerce").fillna(0).sum()
                term_sku_count = term_summary["SKU"].astype(str).replace("", np.nan).dropna().nunique()
                term_table_height = min(360, max(142, 76 + (len(term_summary) * 31)))

                with st.expander(f"{term} | {term_sku_count:,} SKU(s) | Qty In {fmt_num(term_qty_in)} | Qty Out {fmt_num(term_qty_out)}", expanded=len(do_found_terms) <= 5):
                    show_limited_dataframe(term_summary, height=term_table_height, limit=500, show_count=False)

            do_detail_cols = [
                "DO #",
                "Excel Row",
                "SKU",
                "Description",
                "Activity Date",
                "Transaction Type",
                "Trans. #",
                "Qty In",
                "Qty Out",
                "Balance After Transaction",
                "Is Not Shipped",
                "Is Cancelled",
            ]
            do_detail_df = do_detail_df.sort_values(["Searched DO #", "Activity Date", "Excel Row", "SKU"], ascending=[True, False, False, True])
            do_detail_display_df = do_detail_df.rename(columns={"Searched DO #": "DO #"})
            with st.expander("Detailed Matching Transactions", expanded=False):
                st.markdown("<div class='section-subtitle'>Original transaction rows with DO # shown as the first column.</div>", unsafe_allow_html=True)
                show_transaction_dataframe(do_detail_display_df[do_detail_cols], height=380, limit=500)



elif selected_page == "Stock Check":
    tab_page_header("Stock Check", "Enter DO demand to calculate temporary remaining stock. Existing report DOs are recognized to prevent double-counting.")

    stock_table_version_key = f"stock_check_table_version_{site_key}"
    stock_result_signature_key = f"stock_check_result_signature_{site_key}"
    stock_detail_result_key = f"stock_check_detail_result_{site_key}"
    stock_overview_result_key = f"stock_check_overview_result_{site_key}"
    stock_issues_result_key = f"stock_check_issues_result_{site_key}"
    stock_temp_result_key = f"stock_check_temp_balance_result_{site_key}"
    stock_temp_filter_key = f"temp_balance_sku_select_{site_key}"
    stock_saved_table_key = f"stock_check_table_data_{site_key}"
    stock_run_input_key = f"stock_check_run_input_{site_key}"
    stock_rma_excluded_key = f"stock_check_rma_excluded_{site_key}"
    stock_result_model_key = f"stock_check_result_model_{site_key}"
    st.session_state.setdefault(stock_table_version_key, 0)
    stock_table_key = f"stock_check_table_input_{site_key}_{st.session_state[stock_table_version_key]}"
    default_stock_table_df = pd.DataFrame(
        {
            "DO #": [""] * 30,
            "Item Code / SKU": [""] * 30,
            "Qty": [""] * 30,
        }
    )
    saved_stock_table_df = st.session_state.get(stock_saved_table_key)
    if isinstance(saved_stock_table_df, pd.DataFrame) and set(["DO #", "Item Code / SKU", "Qty"]).issubset(saved_stock_table_df.columns):
        stock_template_df = saved_stock_table_df[["DO #", "Item Code / SKU", "Qty"]].copy()
    else:
        stock_template_df = default_stock_table_df
    with st.form(key=f"stock_check_form_{site_key}", clear_on_submit=False):
        reset_col, run_col, action_spacer_col = st.columns([1, 1.5, 4.5])
        with reset_col:
            stock_reset_clicked = st.form_submit_button(
                "Reset",
                use_container_width=True,
            )
        with run_col:
            stock_run_clicked = st.form_submit_button(
                "Run Stock Check",
                type="primary",
                use_container_width=True,
            )
        stock_table_df = st.data_editor(
            stock_template_df,
            use_container_width=True,
            hide_index=True,
            height=342,
            num_rows="dynamic",
            key=stock_table_key,
            column_config={
                "DO #": st.column_config.TextColumn("DO #", help="Paste or enter the DO number."),
                "Item Code / SKU": st.column_config.TextColumn("Item Code / SKU", help="Paste or enter the item code/SKU."),
                "Qty": st.column_config.TextColumn("Qty", help="Paste or enter the requested quantity."),
            },
        )

    if stock_reset_clicked:
        st.session_state[stock_table_version_key] += 1
        keys_to_clear = [stock_result_signature_key, stock_detail_result_key, stock_overview_result_key, stock_issues_result_key, stock_temp_result_key, stock_temp_filter_key, stock_saved_table_key, stock_run_input_key, stock_rma_excluded_key, stock_result_model_key]
        for key in keys_to_clear:
            st.session_state.pop(key, None)
        update_persistent_app_state(remove_keys=keys_to_clear)
        st.rerun()

    row_count_mismatch = False
    stock_has_saved_result = stock_result_signature_key in st.session_state

    if stock_run_clicked:
        st.session_state[stock_saved_table_key] = stock_table_df.copy()
        persistent_stock_values = {stock_saved_table_key: stock_table_df}
        input_df = parse_stock_check_table(stock_table_df)
        if not input_df.empty:
            input_signature_source = input_df.fillna("").astype(str).to_json(orient="records")
            stock_current_signature = hashlib.sha256(f"{uploaded_key}|{input_signature_source}".encode("utf-8")).hexdigest()
            calculation_input_df, excluded_rma_dos = exclude_rma_stock_check_rows(input_df)
            if calculation_input_df.empty:
                detail_df = pd.DataFrame()
                overview_df = pd.DataFrame()
                issues_df = pd.DataFrame()
                temporary_balance_df = pd.DataFrame()
            else:
                if (calculation_input_df["Issue"].astype(str) == "").any():
                    stock_do_cache_key = f"_stock_do_tables_{site_key}"
                    stock_do_cache = st.session_state.get(stock_do_cache_key, {})
                    if isinstance(stock_do_cache, dict) and stock_do_cache.get("source") == model_source_value and isinstance(stock_do_cache.get("tables"), tuple):
                        existing_do_tables = stock_do_cache["tables"]
                    else:
                        existing_do_tables = build_existing_do_tables(model.get("tx_df", pd.DataFrame()))
                        st.session_state[stock_do_cache_key] = {"source": model_source_value, "tables": existing_do_tables}
                else:
                    existing_do_tables = (set(), {}, {})
                detail_df, overview_df, issues_df = build_stock_check_tables(calculation_input_df, sku_df, model.get("tx_df", pd.DataFrame()), existing_do_tables)
                temporary_balance_df = build_temporary_balance_table(sku_df, detail_df) if not detail_df.empty else pd.DataFrame()
            st.session_state[stock_result_signature_key] = stock_current_signature
            st.session_state[stock_detail_result_key] = detail_df
            st.session_state[stock_overview_result_key] = overview_df
            st.session_state[stock_issues_result_key] = issues_df
            st.session_state[stock_temp_result_key] = temporary_balance_df
            st.session_state[stock_run_input_key] = calculation_input_df.copy()
            st.session_state[stock_rma_excluded_key] = excluded_rma_dos
            st.session_state[stock_result_model_key] = model_source_value
            persistent_stock_values.update(
                {
                    stock_result_signature_key: stock_current_signature,
                    stock_detail_result_key: detail_df,
                    stock_overview_result_key: overview_df,
                    stock_issues_result_key: issues_df,
                    stock_temp_result_key: temporary_balance_df,
                    stock_run_input_key: calculation_input_df,
                    stock_rma_excluded_key: excluded_rma_dos,
                    stock_result_model_key: model_source_value,
                }
            )
            stock_has_saved_result = True
        update_persistent_app_state(values=persistent_stock_values)

    if stock_has_saved_result:
        detail_df = st.session_state.get(stock_detail_result_key, pd.DataFrame())
        overview_df = st.session_state.get(stock_overview_result_key, pd.DataFrame())
        issues_df = st.session_state.get(stock_issues_result_key, pd.DataFrame())
        temporary_balance_df = st.session_state.get(stock_temp_result_key, pd.DataFrame())
        result_input_df = st.session_state.get(stock_run_input_key, pd.DataFrame())
        excluded_rma_dos = st.session_state.get(stock_rma_excluded_key, [])
        stock_result_model = st.session_state.get(stock_result_model_key, "")
    else:
        detail_df = pd.DataFrame()
        overview_df = pd.DataFrame()
        issues_df = pd.DataFrame()
        temporary_balance_df = pd.DataFrame()
        result_input_df = pd.DataFrame()
        excluded_rma_dos = []
        stock_result_model = ""

    if stock_has_saved_result and stock_result_model and stock_result_model != model_source_value:
        st.warning(f"Report changed or refreshed. Click Run Stock Check to include all {len(sku_df):,} current SKUs.")

    if stock_has_saved_result and excluded_rma_dos:
        st.warning(f"Excluded {len(excluded_rma_dos):,} RMA DO(s): {', '.join(str(value) for value in excluded_rma_dos)}")

    if stock_has_saved_result and not result_input_df.empty and "Issue" in result_input_df.columns:
        valid_rows = result_input_df[result_input_df["Issue"].astype(str) == ""]
        issue_rows = result_input_df[result_input_df["Issue"].astype(str) != ""]
        input_count_col_1, input_count_col_2, input_count_col_3 = st.columns(3)
        with input_count_col_1:
            st.caption(f"Rows entered: {len(result_input_df):,}")
        with input_count_col_2:
            st.caption(f"Valid rows: {len(valid_rows):,}")
        with input_count_col_3:
            st.caption(f"Input issue rows: {len(issue_rows):,}")

    if not stock_has_saved_result:
        st.info("Paste values under DO #, Item Code / SKU, and Qty, then click Run Stock Check.")
    elif row_count_mismatch:
        pass
    else:
        valid_line_count = len(result_input_df[result_input_df["Issue"].astype(str) == ""]) if not result_input_df.empty and "Issue" in result_input_df.columns else 0
        do_checked_count = overview_df["DO #"].nunique() if not overview_df.empty and "DO #" in overview_df.columns else 0
        existing_do_count = int((overview_df["Report DO Status"] == "Existing DO").sum()) if not overview_df.empty and "Report DO Status" in overview_df.columns else 0
        new_do_count = int((overview_df["Report DO Status"] == "New DO").sum()) if not overview_df.empty and "Report DO Status" in overview_df.columns else 0
        shortage_count = int((detail_df["Status"] == "Shortage").sum()) if not detail_df.empty and "Status" in detail_df.columns else 0
        not_found_count = int((detail_df["Status"] == "Not Found").sum()) if not detail_df.empty and "Status" in detail_df.columns else 0
        already_covered_count = int((detail_df["Status"] == "Already Covered").sum()) if not detail_df.empty and "Status" in detail_df.columns else 0
        data_issue_item_count = int((detail_df["Status"] == "Data Issue").sum()) if not detail_df.empty and "Status" in detail_df.columns else 0
        qty_to_check_total = pd.to_numeric(detail_df["Qty To Check"], errors="coerce").fillna(0).sum() if not detail_df.empty and "Qty To Check" in detail_df.columns else 0
        total_requested_qty = pd.to_numeric(detail_df["Requested Qty"], errors="coerce").fillna(0).sum() if not detail_df.empty and "Requested Qty" in detail_df.columns else 0

        st.markdown("<div class='kpi-row-gap'></div>", unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            metric_card("DOs Checked", fmt_num(do_checked_count), f"Input items: {fmt_num(valid_line_count)}")
        with c2:
            metric_card("Existing / New", f"{fmt_num(existing_do_count)} / {fmt_num(new_do_count)}", "Found in report / new demand")
        with c3:
            metric_card("Qty To Check", fmt_num(qty_to_check_total), f"Requested: {fmt_num(total_requested_qty)} | Covered: {fmt_num(already_covered_count)}")
        with c4:
            metric_card("Shortage / Missing", f"{fmt_num(shortage_count)} / {fmt_num(not_found_count)}", f"Data issues: {fmt_num(data_issue_item_count)}")

        if not overview_df.empty:
            st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>DO Stock Check Overview</div>", unsafe_allow_html=True)
            st.markdown("<div class='section-subtitle'>Each DO # is shown in pasted order. Existing Report Qty Out is treated as already deducted; only Qty To Check is tested against remaining stock.</div>", unsafe_allow_html=True)
            overview_display = prepare_stock_check_display(overview_df)
            overview_height = min(310, max(132, 74 + (len(overview_display) * 32)))
            show_limited_dataframe(overview_display, height=overview_height, limit=1000, show_count=False)

        if not detail_df.empty:
            existing_detail = detail_df[detail_df["Report DO Status"] == "Existing DO"].copy() if "Report DO Status" in detail_df.columns else pd.DataFrame()
            if not existing_detail.empty:
                st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
                st.markdown("<div class='section-title'>Existing DO Found in Report</div>", unsafe_allow_html=True)
                st.markdown("<div class='section-subtitle'>These lines already exist in the report. Only Qty To Check above the existing report quantity is deducted from available stock.</div>", unsafe_allow_html=True)
                existing_cols = ["DO #", "SKU", "Description", "Requested Qty", "Existing Report Qty Out", "Qty To Check", "Status"]
                existing_display = prepare_stock_check_display(existing_detail[existing_cols])
                existing_height = min(320, max(132, 74 + (len(existing_display) * 32)))
                show_limited_dataframe(existing_display, height=existing_height, limit=1000, show_count=False)

            shortage_detail = detail_df[detail_df["Status"].isin(["Shortage", "Not Found"])].copy()
            if not shortage_detail.empty:
                st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
                st.markdown("<div class='section-title'>Items Needing Review</div>", unsafe_allow_html=True)
                st.markdown("<div class='section-subtitle'>Shortage and missing SKU lines are separated here for faster action.</div>", unsafe_allow_html=True)
                shortage_display = prepare_stock_check_display(shortage_detail.drop(columns=["Input Order"], errors="ignore"))
                shortage_height = min(320, max(132, 74 + (len(shortage_display) * 32)))
                show_limited_dataframe(shortage_display, height=shortage_height, limit=1000, show_count=False)

            st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>Item Availability by DO #</div>", unsafe_allow_html=True)
            st.markdown("<div class='section-subtitle'>Current stock is the official Ending Balance. New DOs reduce stock sequentially. Existing DOs only deduct incremental Qty To Check.</div>", unsafe_allow_html=True)

            for do_no in overview_df["DO #"].astype(str).tolist() if not overview_df.empty else []:
                do_items = detail_df[detail_df["DO #"].astype(str) == do_no].copy()
                if do_items.empty:
                    continue
                do_status = overview_df.loc[overview_df["DO #"].astype(str) == do_no, "Status"].iloc[0]
                do_requested = pd.to_numeric(do_items["Requested Qty"], errors="coerce").fillna(0).sum()
                do_qty_to_check = pd.to_numeric(do_items["Qty To Check"], errors="coerce").fillna(0).sum()
                do_shortage = pd.to_numeric(do_items["Shortage Qty"], errors="coerce").fillna(0).sum()
                do_report_status = overview_df.loc[overview_df["DO #"].astype(str) == do_no, "Report DO Status"].iloc[0] if "Report DO Status" in overview_df.columns else "New DO"
                expander_label = f"{do_no} | {do_report_status} | {stock_status_badge(do_status)} | Items {len(do_items):,} | To Check {fmt_num(do_qty_to_check)} | Shortage {fmt_num(do_shortage)}"
                with st.expander(expander_label, expanded=(len(overview_df) <= 4 or do_status not in ["Enough", "Already Covered"])):
                    if do_report_status == "Existing DO":
                        st.markdown(f"<div class='stock-do-subtitle'><b>{html.escape(do_no)}</b> already exists in the report. Existing report qty is not deducted again; only Qty To Check affects remaining stock.</div>", unsafe_allow_html=True)
                    else:
                        st.markdown(f"<div class='stock-do-subtitle'><b>{html.escape(do_no)}</b> is treated as new demand and deducted sequentially after prior pasted DO lines.</div>", unsafe_allow_html=True)
                    display_cols = [
                        "DO #",
                        "SKU",
                        "Description",
                        "Report DO Status",
                        "Report Item Status",
                        "Current Stock",
                        "Available Before",
                        "Requested Qty",
                        "Existing Report Qty Out",
                        "Qty To Check",
                        "Remaining After This Check",
                        "Shortage Qty",
                        "Status",
                    ]
                    do_display = prepare_stock_check_display(do_items[display_cols])
                    do_height = min(360, max(142, 76 + (len(do_display) * 31)))
                    show_limited_dataframe(do_display, height=do_height, limit=500, show_count=False)

            if not temporary_balance_df.empty:
                affected_sku_count = int((pd.to_numeric(temporary_balance_df["Total Qty To Check"], errors="coerce").fillna(0) > 0).sum())
                temporary_shortage_count = int((temporary_balance_df["Temporary Status"] == "Shortage").sum()) if "Temporary Status" in temporary_balance_df.columns else 0
                st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
                st.markdown("<div class='section-title'>Temporary Balance by SKU</div>", unsafe_allow_html=True)
                st.markdown(f"<div class='section-subtitle'>All SKUs are included. Affected SKUs are shown first based on the pasted DO demand. Affected SKUs: {affected_sku_count:,} | Temporary shortage SKUs: {temporary_shortage_count:,}</div>", unsafe_allow_html=True)
                temp_sku_options = ["All SKUs"] + temporary_balance_df["SKU"].astype(str).dropna().tolist()
                if st.session_state.get(stock_temp_filter_key) not in temp_sku_options:
                    st.session_state[stock_temp_filter_key] = "All SKUs"
                    update_persistent_app_state(values={stock_temp_filter_key: "All SKUs"})
                temp_sku_filter = st.selectbox(
                    "Choose SKU",
                    options=temp_sku_options,
                    index=temp_sku_options.index(st.session_state.get(stock_temp_filter_key, "All SKUs")),
                    key=stock_temp_filter_key,
                )
                update_persistent_app_state(values={stock_temp_filter_key: temp_sku_filter})
                filtered_temp_balance_df = temporary_balance_df.copy()
                if temp_sku_filter != "All SKUs":
                    filtered_temp_balance_df = filtered_temp_balance_df[filtered_temp_balance_df["SKU"].astype(str) == str(temp_sku_filter)].copy()
                temp_display = filtered_temp_balance_df.drop(columns=["Impact Sort", "Status Sort"], errors="ignore")
                temp_height = min(560, max(190, 76 + (min(len(temp_display), 14) * 31)))
                show_temporary_balance_dataframe(temp_display, height=temp_height, limit=2000, show_count=True)

            download_col_1, download_col_2, download_col_3 = st.columns(3)
            with download_col_1:
                st.download_button(
                    "Export detail CSV",
                    data=lambda export_df=detail_df: prepare_stock_check_display(export_df.drop(columns=["Input Order"], errors="ignore")).to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"{safe_format_slug(format_name)}_Stock_Check_Detail_{pd.to_datetime(report_end).strftime('%m%d%Y')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    on_click="ignore",
                )
            with download_col_2:
                st.download_button(
                    "Export overview CSV",
                    data=lambda export_df=overview_df: prepare_stock_check_display(export_df).to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"{safe_format_slug(format_name)}_Stock_Check_Overview_{pd.to_datetime(report_end).strftime('%m%d%Y')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    on_click="ignore",
                )
            with download_col_3:
                if not temporary_balance_df.empty:
                    st.download_button(
                        "Export temporary balance CSV",
                        data=lambda export_df=temporary_balance_df: prepare_stock_check_display(export_df.drop(columns=["Impact Sort", "Status Sort"], errors="ignore")).to_csv(index=False).encode("utf-8-sig"),
                        file_name=f"{safe_format_slug(format_name)}_Temporary_Balance_{pd.to_datetime(report_end).strftime('%m%d%Y')}.csv",
                        mime="text/csv",
                        use_container_width=True,
                        on_click="ignore",
                    )

        if not issues_df.empty:
            with st.expander("Input Issues", expanded=detail_df.empty):
                issues_display = issues_df.drop(columns=["Input Order"], errors="ignore").copy()
                if "Requested Qty" in issues_display.columns:
                    issues_display["Requested Qty"] = pd.to_numeric(issues_display["Requested Qty"], errors="coerce")
                issue_height = min(280, max(120, 74 + (len(issues_display) * 32)))
                show_limited_dataframe(issues_display, height=issue_height, limit=1000, show_count=False)

elif selected_page == "Audit":
    tab_page_header("Audit Checks", "Review source completeness, official-row reconciliation, and recent outbound windows.")

    audit_df = model["audit_df"]
    audit_review_df = audit_df[audit_df["Audit Status"] != "Pass"]
    if audit_review_df.empty:
        st.success("All available audit checks passed.")
    else:
        st.markdown('<div class="section-title">Items requiring review</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-subtitle">Only audit exceptions are shown below.</div>', unsafe_allow_html=True)
        show_limited_dataframe(audit_review_df, height=320, limit=500)

    with st.expander("Full SKU Reconciliation", expanded=False):
        show_limited_dataframe(audit_df, height=360, limit=1000)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Outbound window audit</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">Totals are calculated from dated outbound transaction rows.</div>', unsafe_allow_html=True)
    r1, r2, r3, r4 = st.columns(4)
    recent_labels = ["Outbound Last 90 Days", "Outbound Last 30 Days", "Outbound Last 14 Days", "Outbound Last 7 Days"]
    recent_card_labels = ["Recent Outbound 90D", "Recent Outbound 30D", "Recent Outbound 14D", "Recent Outbound 7D"]
    for col, label, card_label in zip([r1, r2, r3, r4], recent_labels, recent_card_labels):
        start, end = windows[label]
        valid_dates = model["window_dates"][label]
        with col:
            metric_card(card_label, fmt_num(sku_df[label].sum()), f"{fmt_date(start)} - {fmt_date(end)} | {len(valid_dates)} data dates")

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Official total rows</div>', unsafe_allow_html=True)
    show_limited_dataframe(model["official_total_df"], height=260, limit=250)

    st.markdown('<div class="section-title" style="margin-top:14px;">Official ending balance rows</div>', unsafe_allow_html=True)
    show_limited_dataframe(model["official_ending_df"], height=260, limit=250)

    with st.expander("Not Shipped Rows", expanded=False):
        show_limited_dataframe(model["not_shipped_df"], height=260, limit=250)

    with st.expander("Cancelled Rows", expanded=False):
        show_limited_dataframe(model["cancelled_df"], height=260, limit=250)

    if not model["beginning_balance_df"].empty:
        with st.expander("Beginning Balance Rows", expanded=False):
            show_limited_dataframe(model["beginning_balance_df"], height=260, limit=250)

elif selected_page == "Help":
    tab_page_header("Help", "The shortest path through the dashboard for daily inventory work.")
    st.markdown(
        """
        1. Select the **Warehouse** and upload the matching Item Activity Report.
        2. Start in **Overview** and review Critical, Warning, and Watch items.
        3. Use **SKU Detail** for item-level activity and **DO Lookup** for order searches.
        4. Use **Stock Check** before creating outbound orders, then use **Audit** only when source reconciliation is needed.
        5. Forecast dates exclude weekends and U.S. federal holidays.
        """
    )
if show_upload_effect:
    sku_count = len(model["sku_df"])
    report_end_text = fmt_date(model.get("report_end"))
    status_box.markdown(
        f"""
        <div class="ready-stage-card">
            <div class="ready-row">
                <div class="ready-check">✓</div>
                <div class="stage-copy">
                    <div class="stage-title">Report ready</div>
                    <div class="stage-subtitle">{sku_count:,} SKUs · {html.escape(format_name)} · {html.escape(selected_page)} ready · Report through {html.escape(report_end_text)}</div>
                    <div class="stage-meta">{html.escape(active_file_name)}</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.session_state["last_upload_effect_key"] = uploaded_key
    st.toast("Report loaded successfully.")
